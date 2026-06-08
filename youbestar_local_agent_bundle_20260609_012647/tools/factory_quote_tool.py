from __future__ import annotations

import hashlib
import json
import math
import re
import sqlite3
from contextlib import contextmanager
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any


DEFAULT_SOURCE_PATH = Path(r"D:\工厂报价")
DATA_DIR = Path("data")
DEFAULT_LIBRARY_PATH = DATA_DIR / "factory_quotes.sqlite3"
IMAGE_TYPE_SKU = "sku_image"
IMAGE_TYPE_REAL = "real_photo"
PHONE_RE = re.compile(r"(?:\+?86[-\s]?)?(1[3-9]\d[-\s]?\d{4}[-\s]?\d{4}|\d{3,4}[-\s]?\d{7,8}(?:[-转]\d{1,6})?)")
DEFAULT_COLUMNS = [
    "SKU图",
    "货号",
    "品名",
    "系列",
    "装箱数",
    "成本单价(元)",
    "成本箱价(元)",
    "产品尺寸(cm)",
    "包装尺寸(cm)",
    "箱规尺寸(cm)",
    "箱毛重(kg)",
    "箱净重(kg)",
    "单品毛重(g)",
    "单品净重(g)",
    "快递包装重量(g)",
    "含税",
    "含运费",
]


HEADER_ALIASES = {
    "category": ("中文包装", "系列", "分类"),
    "image": ("产品图片", "图片"),
    "sku": ("货号", "款号", "型号"),
    "product_name": ("品名", "产品名称", "名称"),
    "package_type": ("包装", "包装方式"),
    "pcs_per_carton": ("装箱数量", "装箱数", "每箱数量"),
    "cost_unit_price": ("品牌价", "单价", "成本价", "出厂价"),
    "cost_carton_price": ("单价    (元/件)", "单价(元/件)", "箱价", "每箱价格"),
    "carton_size_cm": ("外箱规格(cm)", "外箱规格", "外箱尺寸"),
    "package_size_cm": ("包装规格(cm)", "包装规格", "包装尺寸"),
    "product_size_cm": ("产品尺寸（cm）", "产品尺寸(cm)", "产品尺寸"),
    "gross_weight_kg": ("毛重\n(公斤)", "毛重(公斤)", "毛重"),
    "net_weight_kg": ("净重\n(公斤)", "净重(公斤)", "净重"),
    "barcode": ("条码", "商品条码"),
    "notes": ("备注", "说明"),
}


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _clean_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def _clean_phone(value: Any) -> str:
    return re.sub(r"(?<=\d)[\s-]+(?=\d)", "", _clean_text(value))


def _normalize_sku(value: Any) -> str:
    return _clean_text(value).upper()


def _safe_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _decimal_from_value(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _safe_int(value: Any) -> int | None:
    number = _safe_float(value)
    if number is None:
        return None
    return int(number)


def _format_number(value: Any) -> str:
    number = _safe_float(value)
    if number is None:
        return "-" if value in (None, "") else str(value)
    return ("%.4f" % number).rstrip("0").rstrip(".")


def _format_price(value: Any) -> str:
    number = _decimal_from_value(value)
    if number is None:
        return "-" if value in (None, "") else str(value)
    return str(number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _round_price(value: Any) -> float:
    number = _decimal_from_value(value) or Decimal("0")
    return float(number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _format_currency(value: Any) -> str:
    price = _format_price(value)
    return "-" if price == "-" else f"{price}元"


def _format_kg_with_unit(value: Any) -> str:
    weight = _format_number(value)
    return "-" if weight == "-" else f"{weight}kg"


def _format_grams(value: Any) -> str:
    number = _safe_float(value)
    if number is None:
        return "-" if value in (None, "") else str(value)
    return ("%.1f" % number).rstrip("0").rstrip(".")


def _format_grams_with_unit(value: Any, approximate: bool = False, source_note: str = "") -> str:
    grams = _format_grams(value)
    if grams == "-":
        return "未记录"
    prefix = "约" if approximate else ""
    suffix = f"（{source_note}）" if source_note else ""
    return f"{prefix}{grams}g{suffix}"


def _normalize_dimension_cm(value: Any) -> str:
    text = _clean_text(value)
    return re.sub(r"\s*(?:cm|厘米|公分)\s*$", "", text, flags=re.I).strip()


def _format_cm_with_unit(value: Any) -> str:
    size = _normalize_dimension_cm(value)
    return f"{size}cm" if size else "-"


def _parse_weight_to_g(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_text(value).lower()
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group(0))
    if "kg" in text or "公斤" in text or "千克" in text:
        return number * 1000
    return number


def _normalize_image_type(value: Any) -> str:
    raw = _clean_key(value)
    if "实拍" in raw or raw in {"realphoto", "real_photo", "photo", "photos", "actualphoto", "scenephoto"}:
        return IMAGE_TYPE_REAL
    return IMAGE_TYPE_SKU


def _image_type_label(image_type: Any) -> str:
    return "实拍图" if _normalize_image_type(image_type) == IMAGE_TYPE_REAL else "SKU图"


def _stable_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(_clean_text(part).lower() for part in parts)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return f"{prefix}_{digest}"


def _content_hash(data: dict[str, Any]) -> str:
    keys = (
        "factory_name",
        "sku",
        "product_name",
        "package_type",
        "pcs_per_carton",
        "cost_unit_price",
        "carton_size_cm",
        "package_size_cm",
        "product_size_cm",
    )
    raw = "|".join(_clean_text(data.get(key)).lower() for key in keys)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _product_manual_specs(product: dict[str, Any]) -> dict[str, Any]:
    specs = product.get("manual_specs")
    if isinstance(specs, dict):
        return specs
    return {}


def _derived_single_weight_g(product: dict[str, Any], carton_weight_key: str) -> float | None:
    carton_weight_kg = _safe_float(product.get(carton_weight_key))
    pcs_per_carton = _safe_int(product.get("pcs_per_carton"))
    if carton_weight_kg is None or not pcs_per_carton:
        return None
    return round(carton_weight_kg * 1000 / pcs_per_carton, 1)


def _product_weight_profile(product: dict[str, Any]) -> dict[str, Any]:
    manual_specs = _product_manual_specs(product)
    manual_net = _parse_weight_to_g(manual_specs.get("single_net_weight_g"))
    manual_gross = _parse_weight_to_g(manual_specs.get("single_gross_weight_g"))
    shipping_packaged = _parse_weight_to_g(manual_specs.get("shipping_packaged_weight_g"))
    derived_net = _derived_single_weight_g(product, "net_weight_kg")
    derived_gross = _derived_single_weight_g(product, "gross_weight_kg")
    return {
        "single_net_weight_g": manual_net if manual_net is not None else derived_net,
        "single_net_weight_source": "manual" if manual_net is not None else ("carton_derived" if derived_net is not None else "missing"),
        "single_gross_weight_g": manual_gross if manual_gross is not None else derived_gross,
        "single_gross_weight_source": "manual" if manual_gross is not None else ("carton_derived" if derived_gross is not None else "missing"),
        "shipping_packaged_weight_g": shipping_packaged,
        "shipping_packaged_weight_source": "manual" if shipping_packaged is not None else "missing",
    }


def _factory_contact(product: dict[str, Any]) -> dict[str, Any]:
    contact = product.get("factory_contact")
    if not isinstance(contact, dict):
        return {"business_contact": "", "business_phone": "", "raw": ""}
    return {
        "business_contact": _clean_text(contact.get("business_contact")),
        "business_phone": _clean_phone(contact.get("business_phone")),
        "raw": _clean_text(contact.get("raw")),
    }


def _single_weight_text(profile: dict[str, Any], key: str, source_key: str) -> str:
    source = profile.get(source_key)
    return _format_grams_with_unit(
        profile.get(key),
        approximate=source == "carton_derived",
        source_note="箱重/装箱数换算" if source == "carton_derived" else "",
    )


def _product_data_json(product: dict[str, Any]) -> str:
    return json.dumps(product, ensure_ascii=False, sort_keys=True)


def _resolve_source_paths(source_path: str | None = None, workbook_path: str | None = None) -> list[Path]:
    raw_path = workbook_path or source_path
    path = Path(raw_path) if raw_path else DEFAULT_SOURCE_PATH
    if path.is_file():
        return [path]
    if not path.exists():
        return []
    return sorted(item for item in path.glob("*.xlsx") if not item.name.startswith("~$"))


def _library_path(params: dict[str, Any] | None = None) -> Path:
    params = params or {}
    return Path(str(params.get("library_path") or DEFAULT_LIBRARY_PATH))


def _source_mtime(path: str) -> str:
    try:
        timestamp = Path(path).stat().st_mtime
    except OSError:
        return ""
    return datetime.fromtimestamp(timestamp, timezone.utc).isoformat()


@contextmanager
def _connect_library(library_path: Path):
    library_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(library_path)
    connection.row_factory = sqlite3.Row
    try:
        _ensure_library_schema(connection)
        yield connection
    finally:
        connection.close()


def _ensure_library_schema(connection: sqlite3.Connection) -> None:
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS factory_products (
            id TEXT PRIMARY KEY,
            factory_id TEXT NOT NULL,
            factory_name TEXT NOT NULL,
            sku TEXT NOT NULL,
            sku_normalized TEXT NOT NULL,
            product_name TEXT NOT NULL,
            series TEXT,
            barcode TEXT,
            content_hash TEXT NOT NULL,
            source_path TEXT NOT NULL,
            source_updated_at TEXT,
            imported_at TEXT NOT NULL,
            data_json TEXT NOT NULL
        )
        """
    )
    connection.execute("CREATE INDEX IF NOT EXISTS idx_factory_products_factory ON factory_products(factory_name)")
    connection.execute("CREATE INDEX IF NOT EXISTS idx_factory_products_sku ON factory_products(sku_normalized)")
    connection.execute("CREATE INDEX IF NOT EXISTS idx_factory_products_content_hash ON factory_products(content_hash)")
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS image_assets (
            id TEXT PRIMARY KEY,
            product_quote_id TEXT,
            factory_name TEXT,
            sku TEXT,
            source_type TEXT,
            source_url TEXT,
            original_image_url TEXT,
            managed_url TEXT,
            image_type TEXT,
            status TEXT NOT NULL,
            data_json TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    connection.commit()


def _image_url(asset: dict[str, Any]) -> str:
    return _clean_text(asset.get("thumbnail_url") or asset.get("managed_url") or asset.get("original_image_url") or asset.get("source_url"))


def _image_markdown(asset: dict[str, Any] | None, alt: str = "SKU图") -> str:
    if not asset:
        return "未绑定"
    url = _image_url(asset)
    if not url:
        return "未绑定"
    return f"![{alt}]({url})"


def _load_openpyxl():
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法读取 Excel 报价表。") from exc
    return openpyxl


def _detect_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 20) + 1):
        values = [_clean_key(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if "货号" in values and ("品名" in values or "产品名称" in values):
            return row
    return 3


def _header_map(ws, header_row: int) -> dict[str, int]:
    raw_headers = {
        _clean_key(ws.cell(header_row, col).value): col
        for col in range(1, ws.max_column + 1)
        if _clean_key(ws.cell(header_row, col).value)
    }
    mapped: dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            col = raw_headers.get(_clean_key(alias))
            if col:
                mapped[canonical] = col
                break
    return mapped


def _factory_name_from_sheet(ws, fallback: str) -> str:
    for row in range(1, min(ws.max_row, 3) + 1):
        for col in range(1, min(ws.max_column, 5) + 1):
            value = ws.cell(row, col).value
            if not value:
                continue
            first_line = str(value).splitlines()[0].strip()
            if first_line:
                return first_line
    return fallback


def _header_metadata_text(ws, header_row: int) -> str:
    rows: list[str] = []
    max_row = min(ws.max_row, max(3, header_row))
    for row in range(1, max_row + 1):
        values: list[str] = []
        for col in range(1, ws.max_column + 1):
            value = _clean_text(ws.cell(row, col).value)
            if value:
                values.append(value)
        if values:
            rows.append(" ".join(values))
    return "\n".join(rows)


def _raw_contact_line(text: str) -> str:
    for line in str(text or "").splitlines():
        if re.search(r"业务员|业务联系人|业务联系|联系人|联系电话|电话|手机", line, re.I):
            return _clean_text(line)
    phone_match = PHONE_RE.search(text or "")
    return _clean_text(phone_match.group(0)) if phone_match else ""


def _extract_business_phone(text: str) -> str:
    match = PHONE_RE.search(text or "")
    return _clean_phone(match.group(0)) if match else ""


def _extract_business_contact(text: str) -> str:
    match = re.search(r"(?:业务员|业务联系人|业务联系|联系人|销售)\s*[:：]?\s*([^\n\r；;]+)", text or "", re.I)
    if not match:
        return ""
    candidate = match.group(1)
    candidate = re.split(r"(?:联系电话|电话|手机|tel|mobile)\s*[:：]?", candidate, maxsplit=1, flags=re.I)[0]
    candidate = PHONE_RE.sub("", candidate)
    candidate = re.sub(r"^[：:\s,，;；]+|[：:\s,，;；。]+$", "", candidate)
    return _clean_text(candidate)


def _factory_metadata_from_sheet(ws, fallback: str, header_row: int) -> dict[str, str]:
    header_text = _header_metadata_text(ws, header_row)
    return {
        "factory_name": _factory_name_from_sheet(ws, fallback),
        "business_contact": _extract_business_contact(header_text),
        "business_phone": _extract_business_phone(header_text),
        "raw": _raw_contact_line(header_text),
    }


def _cell_value(ws, data_ws, row: int, col: int | None) -> Any:
    if not col:
        return None
    value = data_ws.cell(row, col).value
    if value is None:
        value = ws.cell(row, col).value
    return value


def _build_product(
    row: int,
    ws,
    data_ws,
    columns: dict[str, int],
    factory_metadata: dict[str, str],
    source_file: Path,
) -> dict[str, Any] | None:
    factory_name = factory_metadata["factory_name"]
    sku = _normalize_sku(_cell_value(ws, data_ws, row, columns.get("sku")))
    product_name = _clean_text(_cell_value(ws, data_ws, row, columns.get("product_name")))
    if not sku and not product_name:
        return None

    pcs_per_carton = _safe_int(_cell_value(ws, data_ws, row, columns.get("pcs_per_carton")))
    unit_price = _safe_float(_cell_value(ws, data_ws, row, columns.get("cost_unit_price")))
    carton_price = _safe_float(_cell_value(ws, data_ws, row, columns.get("cost_carton_price")))
    if carton_price is None and unit_price is not None and pcs_per_carton:
        carton_price = round(unit_price * pcs_per_carton, 4)

    record = {
        "id": _stable_id("fpq", factory_name, sku, product_name, _cell_value(ws, data_ws, row, columns.get("package_type"))),
        "factory_id": _stable_id("factory", factory_name),
        "factory_name": factory_name,
        "factory_contact": {
            "business_contact": factory_metadata.get("business_contact", ""),
            "business_phone": factory_metadata.get("business_phone", ""),
            "raw": factory_metadata.get("raw", ""),
        },
        "sku": sku,
        "product_name": product_name,
        "series": _clean_text(_cell_value(ws, data_ws, row, columns.get("category"))),
        "package_type": _clean_text(_cell_value(ws, data_ws, row, columns.get("package_type"))),
        "pcs_per_carton": pcs_per_carton,
        "cost_unit_price": unit_price,
        "cost_carton_price": carton_price,
        "carton_size_cm": _clean_text(_cell_value(ws, data_ws, row, columns.get("carton_size_cm"))),
        "package_size_cm": _clean_text(_cell_value(ws, data_ws, row, columns.get("package_size_cm"))),
        "product_size_cm": _clean_text(_cell_value(ws, data_ws, row, columns.get("product_size_cm"))),
        "gross_weight_kg": _safe_float(_cell_value(ws, data_ws, row, columns.get("gross_weight_kg"))),
        "net_weight_kg": _safe_float(_cell_value(ws, data_ws, row, columns.get("net_weight_kg"))),
        "barcode": _clean_text(_cell_value(ws, data_ws, row, columns.get("barcode"))),
        "notes": _clean_text(_cell_value(ws, data_ws, row, columns.get("notes"))),
        "image_asset_ids": [],
        "quote_terms": {
            "tax_included": None,
            "freight_included": None,
        },
        "source": {
            "type": "excel",
            "path": str(source_file),
            "sheet": ws.title,
            "row": row,
        },
    }
    record["content_hash"] = _content_hash(record)
    return record


def load_quote_products(source_path: str | None = None, workbook_path: str | None = None) -> list[dict[str, Any]]:
    openpyxl = _load_openpyxl()
    products: list[dict[str, Any]] = []
    for path in _resolve_source_paths(source_path, workbook_path):
        workbook = openpyxl.load_workbook(path, data_only=False)
        data_workbook = openpyxl.load_workbook(path, data_only=True)
        for ws in workbook.worksheets:
            data_ws = data_workbook[ws.title]
            header_row = _detect_header_row(ws)
            columns = _header_map(ws, header_row)
            if "sku" not in columns and "product_name" not in columns:
                continue
            factory_metadata = _factory_metadata_from_sheet(ws, path.stem, header_row)
            current_series = ""
            for row in range(header_row + 1, ws.max_row + 1):
                series_value = _clean_text(_cell_value(ws, data_ws, row, columns.get("category")))
                if series_value:
                    current_series = series_value
                product = _build_product(row, ws, data_ws, columns, factory_metadata, path)
                if not product:
                    continue
                if not product["series"]:
                    product["series"] = current_series
                products.append(product)
    return products


def save_products_to_library(products: list[dict[str, Any]], library_path: str | Path | None = None, replace_source: bool = True) -> dict[str, Any]:
    path = Path(library_path) if library_path else DEFAULT_LIBRARY_PATH
    imported_at = _utc_now()
    source_paths = sorted({_clean_text(product.get("source", {}).get("path")) for product in products if product.get("source")})
    with _connect_library(path) as connection:
        if replace_source:
            for source_path in source_paths:
                connection.execute("DELETE FROM factory_products WHERE source_path = ?", (source_path,))
        for product in products:
            source = product.get("source") if isinstance(product.get("source"), dict) else {}
            source_path = _clean_text(source.get("path"))
            connection.execute(
                """
                INSERT OR REPLACE INTO factory_products (
                    id,
                    factory_id,
                    factory_name,
                    sku,
                    sku_normalized,
                    product_name,
                    series,
                    barcode,
                    content_hash,
                    source_path,
                    source_updated_at,
                    imported_at,
                    data_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    product.get("id"),
                    product.get("factory_id"),
                    product.get("factory_name"),
                    product.get("sku"),
                    _normalize_sku(product.get("sku")),
                    product.get("product_name"),
                    product.get("series"),
                    product.get("barcode"),
                    product.get("content_hash"),
                    source_path,
                    _source_mtime(source_path),
                    imported_at,
                    _product_data_json(product),
                ),
            )
        connection.commit()
    return {"library_path": str(path), "stored_count": len(products), "source_count": len(source_paths), "imported_at": imported_at}


def _product_from_row(row: sqlite3.Row) -> dict[str, Any]:
    data = json.loads(row["data_json"])
    if isinstance(data, dict):
        data.setdefault("id", row["id"])
        return data
    return {}


def load_products_from_library(params: dict[str, Any]) -> list[dict[str, Any]]:
    path = _library_path(params)
    if not path.exists():
        return []
    factory_name = str(params.get("factory_name") or params.get("factory") or "").strip()
    sku = _normalize_sku(params.get("sku") or params.get("货号"))
    keyword = str(params.get("keyword") or params.get("query") or params.get("product_name") or "").strip()

    where = []
    values: list[Any] = []
    if sku:
        where.append("sku_normalized = ?")
        values.append(sku)
    if factory_name:
        where.append("factory_name LIKE ?")
        values.append(f"%{factory_name}%")
    if keyword and not sku:
        where.append("(sku_normalized LIKE ? OR product_name LIKE ? OR series LIKE ? OR barcode LIKE ?)")
        like = f"%{keyword}%"
        values.extend([like, like, like, like])

    sql = "SELECT * FROM factory_products"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY factory_name, sku_normalized, product_name LIMIT 200"

    with _connect_library(path) as connection:
        return [_product_from_row(row) for row in connection.execute(sql, values).fetchall()]


def _asset_from_row(row: sqlite3.Row) -> dict[str, Any]:
    data = json.loads(row["data_json"])
    if isinstance(data, dict):
        data.setdefault("id", row["id"])
        data["status"] = row["status"]
        data["image_type"] = _normalize_image_type(data.get("image_type") or row["image_type"])
        data.setdefault("product_quote_id", row["product_quote_id"] or "")
        data.setdefault("factory_name", row["factory_name"] or "")
        data.setdefault("sku", row["sku"] or "")
        return data
    return {}


def load_image_assets_for_product(
    product: dict[str, Any],
    params: dict[str, Any] | None = None,
    status: str = "confirmed",
    image_type: str | None = None,
) -> list[dict[str, Any]]:
    path = _library_path(params)
    if not path.exists():
        return []
    where = ["status = ?"]
    values: list[Any] = [status]
    product_id = _clean_text(product.get("id"))
    factory_name = _clean_text(product.get("factory_name"))
    sku = _normalize_sku(product.get("sku"))
    if product_id:
        where.append("(product_quote_id = ? OR (sku = ? AND (factory_name = ? OR factory_name = '')))")
        values.extend([product_id, sku, factory_name])
    else:
        where.append("sku = ?")
        values.append(sku)
        if factory_name:
            where.append("(factory_name = ? OR factory_name = '')")
            values.append(factory_name)
    sql = "SELECT * FROM image_assets WHERE " + " AND ".join(where) + " ORDER BY updated_at DESC LIMIT 50"
    with _connect_library(path) as connection:
        assets = [_asset_from_row(row) for row in connection.execute(sql, values).fetchall()]
    if image_type:
        normalized_type = _normalize_image_type(image_type)
        assets = [asset for asset in assets if _normalize_image_type(asset.get("image_type")) == normalized_type]
    return assets


def _first_image_asset(product: dict[str, Any], params: dict[str, Any] | None = None) -> dict[str, Any] | None:
    assets = load_image_assets_for_product(product, params, status="confirmed", image_type=IMAGE_TYPE_SKU)
    return assets[0] if assets else None


def _real_photo_assets(product: dict[str, Any], params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    return load_image_assets_for_product(product, params, status="confirmed", image_type=IMAGE_TYPE_REAL)


def _replace_existing_sku_images(connection: sqlite3.Connection, saved: dict[str, Any], now: str) -> None:
    product_id = _clean_text(saved.get("product_quote_id"))
    sku = _normalize_sku(saved.get("sku"))
    factory_name = _clean_text(saved.get("factory_name"))
    if not product_id and not sku:
        return

    where = ["status = ?", "id <> ?"]
    values: list[Any] = ["confirmed", saved["id"]]
    if product_id:
        where.append("(product_quote_id = ? OR (sku = ? AND (factory_name = ? OR factory_name = '')))")
        values.extend([product_id, sku, factory_name])
    else:
        where.append("sku = ?")
        values.append(sku)
        if factory_name:
            where.append("(factory_name = ? OR factory_name = '')")
            values.append(factory_name)

    rows = connection.execute("SELECT * FROM image_assets WHERE " + " AND ".join(where), values).fetchall()
    for row in rows:
        asset = _asset_from_row(row)
        if _normalize_image_type(asset.get("image_type")) != IMAGE_TYPE_SKU:
            continue
        asset["status"] = "replaced"
        asset["replaced_by"] = saved["id"]
        connection.execute(
            "UPDATE image_assets SET status = ?, data_json = ?, updated_at = ? WHERE id = ?",
            ("replaced", json.dumps(asset, ensure_ascii=False, sort_keys=True), now, row["id"]),
        )


def _save_image_asset(candidate: dict[str, Any], params: dict[str, Any], status: str) -> dict[str, Any]:
    saved = dict(candidate)
    saved["status"] = status
    saved["image_type"] = _normalize_image_type(saved.get("image_type"))
    now = _utc_now()
    with _connect_library(_library_path(params)) as connection:
        if status == "confirmed" and saved["image_type"] == IMAGE_TYPE_SKU:
            _replace_existing_sku_images(connection, saved, now)
        connection.execute(
            """
            INSERT OR REPLACE INTO image_assets (
                id,
                product_quote_id,
                factory_name,
                sku,
                source_type,
                source_url,
                original_image_url,
                managed_url,
                image_type,
                status,
                data_json,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, COALESCE((SELECT created_at FROM image_assets WHERE id = ?), ?), ?)
            """,
            (
                saved["id"],
                saved.get("product_quote_id", ""),
                saved.get("factory_name", ""),
                saved.get("sku", ""),
                saved.get("source_type", ""),
                saved.get("source_url", ""),
                saved.get("original_image_url", ""),
                saved.get("managed_url", ""),
                saved.get("image_type", ""),
                saved["status"],
                json.dumps(saved, ensure_ascii=False, sort_keys=True),
                saved["id"],
                now,
                now,
            ),
        )
        connection.commit()
    return saved


def _image_candidate_from_params(params: dict[str, Any], product: dict[str, Any] | None = None) -> dict[str, Any] | None:
    image_url = _clean_text(params.get("image_url") or params.get("url"))
    source_url = _clean_text(params.get("source_url") or params.get("1688_url") or params.get("album_url"))
    source_type = _clean_text(params.get("source_type")) or ("1688_product_page" if "1688.com" in source_url else "external_link")
    sku = _normalize_sku(params.get("sku") or params.get("货号") or (product or {}).get("sku"))
    factory_name = _clean_text(params.get("factory_name") or params.get("factory") or (product or {}).get("factory_name"))
    image_type = _normalize_image_type(params.get("image_type") or params.get("type") or params.get("图片类型"))

    if not image_url and not source_url:
        return None

    return {
        "id": _stable_id("img", factory_name, sku, image_type, image_url or source_url),
        "product_quote_id": _clean_text(params.get("product_quote_id") or (product or {}).get("id")),
        "factory_name": factory_name,
        "sku": sku,
        "source_type": source_type,
        "source_url": source_url,
        "original_image_url": image_url,
        "managed_url": _clean_text(params.get("managed_url")),
        "thumbnail_url": _clean_text(params.get("thumbnail_url") or image_url or source_url),
        "image_type": image_type,
        "status": "pending_confirmation",
    }


def _weight_updates_from_params(params: dict[str, Any]) -> dict[str, float]:
    field_sources = {
        "single_net_weight_g": (
            "single_net_weight_g",
            "single_net_weight",
            "net_weight_g",
            "单品净重",
            "单个产品净重",
            "产品净重",
        ),
        "single_gross_weight_g": (
            "single_gross_weight_g",
            "single_gross_weight",
            "gross_weight_g",
            "单品毛重",
            "单个产品毛重",
            "产品毛重",
        ),
        "shipping_packaged_weight_g": (
            "shipping_packaged_weight_g",
            "shipping_weight_g",
            "packaged_weight_g",
            "包装重量",
            "快递包装重量",
            "打包后重量",
        ),
    }
    updates: dict[str, float] = {}
    for target, aliases in field_sources.items():
        for alias in aliases:
            if alias not in params:
                continue
            grams = _parse_weight_to_g(params.get(alias))
            if grams is not None:
                updates[target] = round(grams, 1)
                break
    return updates


def _spec_updates_from_params(params: dict[str, Any]) -> dict[str, str]:
    field_sources = {
        "product_size_cm": (
            "product_size_cm",
            "product_size",
            "产品尺寸",
            "产品尺寸cm",
        ),
        "package_size_cm": (
            "package_size_cm",
            "package_size",
            "包装尺寸",
            "包装规格",
            "包装尺寸cm",
        ),
        "carton_size_cm": (
            "carton_size_cm",
            "carton_size",
            "box_size_cm",
            "箱规尺寸",
            "外箱尺寸",
            "外箱规格",
        ),
    }
    updates: dict[str, str] = {}
    for target, aliases in field_sources.items():
        for alias in aliases:
            if alias not in params:
                continue
            value = _normalize_dimension_cm(params.get(alias))
            if value:
                updates[target] = value
                break
    return updates


def _update_product_in_library(product: dict[str, Any], params: dict[str, Any]) -> None:
    with _connect_library(_library_path(params)) as connection:
        connection.execute(
            """
            UPDATE factory_products
            SET data_json = ?
            WHERE id = ?
            """,
            (_product_data_json(product), product.get("id")),
        )
        connection.commit()


def library_status(params: dict[str, Any] | None = None) -> dict[str, Any]:
    params = params or {}
    path = _library_path(params)
    if not path.exists():
        return {
            "ok": True,
            "kind": "factory_quote_library_status",
            "title": "工厂报价资料库状态",
            "columns": ["资料库", "工厂数", "产品数", "状态"],
            "rows": [[str(path), 0, 0, "未创建"]],
            "summary": {"资料库": str(path), "工厂数": 0, "产品数": 0, "状态": "未创建，请先执行导入"},
            "data": {"library_path": str(path), "exists": False},
        }

    with _connect_library(path) as connection:
        product_count = connection.execute("SELECT COUNT(*) FROM factory_products").fetchone()[0]
        factory_count = connection.execute("SELECT COUNT(DISTINCT factory_name) FROM factory_products").fetchone()[0]
        latest_import = connection.execute("SELECT MAX(imported_at) FROM factory_products").fetchone()[0]

    return {
        "ok": True,
        "kind": "factory_quote_library_status",
        "title": "工厂报价资料库状态",
        "columns": ["资料库", "工厂数", "产品数", "最近导入"],
        "rows": [[str(path), factory_count, product_count, latest_import or "-"]],
        "summary": {"资料库": str(path), "工厂数": factory_count, "产品数": product_count, "最近导入": latest_import or "-"},
        "data": {"library_path": str(path), "exists": True, "factory_count": factory_count, "product_count": product_count, "latest_import": latest_import},
    }


def _matches_factory(product: dict[str, Any], factory_name: str) -> bool:
    clean_query = _clean_text(factory_name).lower()
    if not clean_query:
        return True
    factory = _clean_text(product.get("factory_name")).lower()
    factory_id = _clean_text(product.get("factory_id")).lower()
    return clean_query in factory or clean_query == factory_id


def _matches_keyword(product: dict[str, Any], keyword: str) -> bool:
    clean_keyword = _clean_text(keyword).lower()
    if not clean_keyword:
        return True
    fields = (
        product.get("sku"),
        product.get("product_name"),
        product.get("series"),
        product.get("barcode"),
        product.get("notes"),
    )
    return any(clean_keyword in _clean_text(field).lower() for field in fields)


def find_products(params: dict[str, Any]) -> list[dict[str, Any]]:
    use_excel = bool(params.get("use_excel") or params.get("workbook_path") or params.get("source_path"))
    products = [] if use_excel else load_products_from_library(params)
    if not products:
        products = load_quote_products(params.get("source_path"), params.get("workbook_path"))
    factory_name = str(params.get("factory_name") or params.get("factory") or "").strip()
    sku = _normalize_sku(params.get("sku") or params.get("货号"))
    keyword = str(params.get("keyword") or params.get("query") or params.get("product_name") or "").strip()

    matches = []
    for product in products:
        if not _matches_factory(product, factory_name):
            continue
        if sku and product.get("sku") != sku:
            continue
        if not sku and keyword and not _matches_keyword(product, keyword):
            continue
        matches.append(product)
    return matches


def _tier_margin_rate(quantity: int, params: dict[str, Any]) -> float | None:
    tiers = params.get("pricing_tiers")
    if not isinstance(tiers, list):
        return _safe_float(params.get("margin_rate"))
    selected: float | None = None
    selected_min = -1
    for tier in tiers:
        if not isinstance(tier, dict):
            continue
        min_qty = _safe_int(tier.get("min_qty")) or 0
        margin_rate = _safe_float(tier.get("margin_rate"))
        if margin_rate is not None and quantity >= min_qty and min_qty >= selected_min:
            selected = margin_rate
            selected_min = min_qty
    return selected


def calculate_quote(product: dict[str, Any], params: dict[str, Any]) -> dict[str, Any] | None:
    quantity = _safe_int(params.get("quantity") or params.get("qty") or params.get("数量"))
    unit_cost = _safe_float(product.get("cost_unit_price"))
    if not quantity or quantity <= 0 or unit_cost is None:
        return None

    margin_rate = _tier_margin_rate(quantity, params)
    tax_rate = _safe_float(params.get("tax_rate")) or 0
    include_tax = bool(params.get("include_tax", False))
    include_freight = bool(params.get("include_freight", False))
    freight_fee = _safe_float(params.get("freight_fee")) or 0
    pcs_per_carton = _safe_int(product.get("pcs_per_carton")) or 0

    quote_unit_price = unit_cost
    if margin_rate is not None:
        quote_unit_price *= 1 + margin_rate
    if include_tax:
        quote_unit_price *= 1 + tax_rate

    product_total = quote_unit_price * quantity
    freight_total = freight_fee if include_freight else 0
    quote_total = product_total + freight_total

    cartons = math.ceil(quantity / pcs_per_carton) if pcs_per_carton else None
    return {
        "quantity": quantity,
        "pcs_per_carton": pcs_per_carton or None,
        "estimated_cartons": cartons,
        "unit_cost": _round_price(unit_cost),
        "cost_total": _round_price(unit_cost * quantity),
        "margin_rate": margin_rate,
        "tax_rate": tax_rate if include_tax else None,
        "include_tax": include_tax,
        "include_freight": include_freight,
        "freight_fee": _round_price(freight_fee) if include_freight else 0,
        "quote_unit_price": _round_price(quote_unit_price),
        "quote_total": _round_price(quote_total),
    }


def _row_for_product(product: dict[str, Any], calculation: dict[str, Any] | None = None, sku_image: dict[str, Any] | None = None) -> list[Any]:
    tax_text = "未知" if product.get("quote_terms", {}).get("tax_included") is None else str(product["quote_terms"]["tax_included"])
    freight_text = "未知" if product.get("quote_terms", {}).get("freight_included") is None else str(product["quote_terms"]["freight_included"])
    weight_profile = _product_weight_profile(product)
    row = [
        _image_markdown(sku_image, "SKU图"),
        product.get("sku", ""),
        product.get("product_name", ""),
        product.get("series", ""),
        product.get("pcs_per_carton") or "-",
        _format_price(product.get("cost_unit_price")),
        _format_price(product.get("cost_carton_price")),
        product.get("product_size_cm") or "-",
        product.get("package_size_cm") or "-",
        product.get("carton_size_cm") or "-",
        _format_number(product.get("gross_weight_kg")),
        _format_number(product.get("net_weight_kg")),
        _single_weight_text(weight_profile, "single_gross_weight_g", "single_gross_weight_source"),
        _single_weight_text(weight_profile, "single_net_weight_g", "single_net_weight_source"),
        _format_grams_with_unit(weight_profile.get("shipping_packaged_weight_g")),
        tax_text,
        freight_text,
    ]
    if calculation:
        row.extend(
            [
                calculation["quantity"],
                _format_price(calculation["quote_unit_price"]),
                _format_price(calculation["quote_total"]),
            ]
        )
    return row


def _summary_for_product(
    product: dict[str, Any],
    calculation: dict[str, Any] | None = None,
    sku_image: dict[str, Any] | None = None,
    real_photos: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    weight_profile = _product_weight_profile(product)
    contact = _factory_contact(product)
    summary = {
        "工厂": product.get("factory_name"),
        "货号": product.get("sku"),
        "品名": product.get("product_name"),
        "装箱数": product.get("pcs_per_carton"),
        "成本单价": _format_currency(product.get("cost_unit_price")),
        "产品尺寸": _format_cm_with_unit(product.get("product_size_cm")),
        "包装尺寸": _format_cm_with_unit(product.get("package_size_cm")),
        "箱规尺寸": _format_cm_with_unit(product.get("carton_size_cm")),
        "箱毛重": _format_kg_with_unit(product.get("gross_weight_kg")),
        "箱净重": _format_kg_with_unit(product.get("net_weight_kg")),
        "单品毛重": _single_weight_text(weight_profile, "single_gross_weight_g", "single_gross_weight_source"),
        "单品净重": _single_weight_text(weight_profile, "single_net_weight_g", "single_net_weight_source"),
        "快递包装重量": _format_grams_with_unit(weight_profile.get("shipping_packaged_weight_g")),
        "SKU图": "已绑定" if sku_image else "未绑定",
        "实拍图数量": len(real_photos or []),
        "含税": "未知",
        "含运费": "未知",
    }
    if contact["business_contact"]:
        summary["业务联系人"] = contact["business_contact"]
    if contact["business_phone"]:
        summary["业务联系电话"] = contact["business_phone"]
    if calculation:
        summary.update(
            {
                "查询数量": calculation["quantity"],
                "报价单价": _format_currency(calculation["quote_unit_price"]),
                "报价总额": _format_currency(calculation["quote_total"]),
            }
        )
    return summary


def _unique_factory_contacts(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    contacts_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    for product in products:
        contact = _factory_contact(product)
        key = (
            _clean_text(product.get("factory_name")),
            contact["business_contact"],
            contact["business_phone"],
        )
        if key in contacts_by_key:
            contacts_by_key[key]["product_count"] += 1
            continue
        contacts_by_key[key] = {
            "factory_name": key[0],
            "business_contact": key[1],
            "business_phone": key[2],
            "raw": contact["raw"],
            "product_count": 1,
        }
    return list(contacts_by_key.values())


def query_factory_contact(params: dict[str, Any]) -> dict[str, Any]:
    params = params or {}
    try:
        matches = find_products(params)
    except Exception as exc:
        return {
            "ok": False,
            "kind": "factory_contact_error",
            "title": "工厂联系人查询失败",
            "message": str(exc),
        }

    if not matches:
        return {
            "ok": False,
            "kind": "factory_contact_not_found",
            "title": "工厂联系人查询失败",
            "message": "没有找到匹配的工厂联系人资料。请先导入报价表，或提供更准确的工厂名称。",
        }

    contacts = _unique_factory_contacts(matches)
    rows = [
        [
            item["factory_name"] or "-",
            item["business_contact"] or "未记录",
            item["business_phone"] or "未记录",
            item["product_count"],
        ]
        for item in contacts[:20]
    ]
    first = contacts[0] if contacts else {}
    return {
        "ok": True,
        "kind": "factory_contact",
        "title": "工厂联系人查询结果",
        "columns": ["工厂", "业务联系人", "业务联系电话", "匹配产品数"],
        "rows": rows,
        "summary": {
            "工厂": first.get("factory_name") or "-",
            "业务联系人": first.get("business_contact") or "未记录",
            "业务联系电话": first.get("business_phone") or "未记录",
            "匹配联系人数量": len(contacts),
            "资料库": str(_library_path(params)),
        },
        "data": {
            "count": len(contacts),
            "library_path": str(_library_path(params)),
            "contacts": contacts[:20],
            "products": matches[:20],
        },
    }


def query_factory_quote(params: dict[str, Any]) -> dict[str, Any]:
    params = params or {}
    try:
        matches = find_products(params)
    except Exception as exc:
        return {
            "ok": False,
            "kind": "factory_quote_error",
            "title": "工厂报价查询失败",
            "message": str(exc),
        }

    if not matches:
        return {
            "ok": False,
            "kind": "factory_quote_not_found",
            "title": "工厂报价查询失败",
            "message": "没有找到匹配的工厂报价资料。请提供工厂名称 + 货号，或检查报价表路径。",
        }

    limited_matches = matches[:20]
    single_match = len(matches) == 1
    calculations = [calculate_quote(product, params) if single_match else None for product in limited_matches]
    sku_images = [_first_image_asset(product, params) for product in limited_matches]
    real_photo_groups = [_real_photo_assets(product, params) for product in limited_matches]
    columns = list(DEFAULT_COLUMNS)
    if single_match and calculations[0]:
        columns.extend(["查询数量", "报价单价", "报价总额"])

    return {
        "ok": True,
        "kind": "factory_product_quote",
        "title": "工厂报价查询结果",
        "columns": columns,
        "rows": [
            _row_for_product(product, calculation, sku_image)
            for product, calculation, sku_image in zip(limited_matches, calculations, sku_images)
        ],
        "summary": _summary_for_product(limited_matches[0], calculations[0], sku_images[0], real_photo_groups[0])
        if single_match
        else {"匹配数量": len(matches), "提示": "货号可能在多个工厂重复，请补充工厂名称。"},
        "data": {
            "count": len(matches),
            "needs_disambiguation": len(matches) > 1,
            "library_path": str(_library_path(params)),
            "source": "excel" if params.get("source_path") or params.get("workbook_path") or params.get("use_excel") else "library",
            "products": limited_matches,
            "contact": _factory_contact(limited_matches[0]) if single_match else None,
            "sku_images": sku_images,
            "real_photos": real_photo_groups,
            "calculation": calculations[0] if single_match else None,
        },
    }


def import_factory_quotes(params: dict[str, Any]) -> dict[str, Any]:
    params = params or {}
    try:
        products = load_quote_products(params.get("source_path"), params.get("workbook_path"))
        save_result = save_products_to_library(products, _library_path(params), replace_source=bool(params.get("replace_source", True)))
    except Exception as exc:
        return {
            "ok": False,
            "kind": "factory_quote_import_error",
            "title": "工厂报价导入失败",
            "message": str(exc),
        }

    factory_names = sorted({_clean_text(product.get("factory_name")) for product in products if product.get("factory_name")})
    return {
        "ok": True,
        "kind": "factory_quote_import",
        "title": "工厂报价导入结果",
        "columns": ["工厂", "产品数量"],
        "rows": [[factory, sum(1 for product in products if product.get("factory_name") == factory)] for factory in factory_names],
        "summary": {
            "工厂数": len(factory_names),
            "产品数": len(products),
            "资料库": save_result["library_path"],
            "状态": "已写入本地报价资料库，未写入长期记忆",
        },
        "data": {
            "products": products[:20],
            "total_count": len(products),
            "source_path": str(params.get("source_path") or params.get("workbook_path") or DEFAULT_SOURCE_PATH),
            **save_result,
        },
    }


def create_image_asset_candidate(params: dict[str, Any]) -> dict[str, Any]:
    params = params or {}
    candidate = _image_candidate_from_params(params)
    if not candidate:
        return {
            "ok": False,
            "kind": "image_asset_candidate_error",
            "title": "图片资产候选生成失败",
            "message": "请提供图片直链、网络相册链接或 1688 商品链接。",
        }
    candidate = _save_image_asset(candidate, params, "pending_confirmation")
    image_label = _image_type_label(candidate.get("image_type"))
    return {
        "ok": True,
        "kind": "image_asset_candidate",
        "title": "图片资产候选",
        "columns": ["货号", "图片类型", "来源类型", "图片链接", "状态"],
        "rows": [[candidate["sku"], image_label, candidate["source_type"], candidate["original_image_url"] or candidate["source_url"], candidate["status"]]],
        "summary": {
            "工厂": candidate["factory_name"] or "未指定",
            "货号": candidate["sku"] or "未指定",
            "图片类型": image_label,
            "状态": "待确认绑定",
            "资料库": str(_library_path(params)),
        },
        "data": candidate,
    }


def bind_image_asset(params: dict[str, Any]) -> dict[str, Any]:
    params = params or {}
    matches = find_products(params)
    if not matches:
        return {
            "ok": False,
            "kind": "image_binding_error",
            "title": "图片绑定失败",
            "message": "没有找到要绑定图片的产品。请提供工厂名称 + 货号。",
        }
    if len(matches) > 1:
        return {
            "ok": False,
            "kind": "image_binding_ambiguous",
            "title": "图片绑定失败",
            "message": "货号匹配到多个产品，请补充工厂名称后再绑定。",
            "data": {"matches": matches[:20]},
        }

    product = matches[0]
    candidate = _image_candidate_from_params(params, product)
    if not candidate and params.get("image_id"):
        image_id = _clean_text(params.get("image_id"))
        with _connect_library(_library_path(params)) as connection:
            row = connection.execute("SELECT * FROM image_assets WHERE id = ?", (image_id,)).fetchone()
        candidate = _asset_from_row(row) if row else None
        if candidate:
            candidate["product_quote_id"] = product.get("id", "")
            candidate["factory_name"] = product.get("factory_name", "")
            candidate["sku"] = product.get("sku", "")

    if not candidate:
        return {
            "ok": False,
            "kind": "image_binding_error",
            "title": "图片绑定失败",
            "message": "请提供图片链接、相册链接、1688 链接或 image_id。",
        }

    candidate["product_quote_id"] = product.get("id", "")
    candidate["factory_name"] = product.get("factory_name", "")
    candidate["sku"] = product.get("sku", "")
    candidate = _save_image_asset(candidate, params, "confirmed")
    image_label = _image_type_label(candidate.get("image_type"))

    return {
        "ok": True,
        "kind": "image_asset_binding",
        "title": "图片绑定结果",
        "columns": ["货号", "品名", "图片类型", "图片", "状态"],
        "rows": [[product.get("sku", ""), product.get("product_name", ""), image_label, _image_markdown(candidate, image_label), "已绑定"]],
        "summary": {
            "工厂": product.get("factory_name"),
            "货号": product.get("sku"),
            "品名": product.get("product_name"),
            "图片类型": image_label,
            "状态": "已绑定",
            "图片": _image_markdown(candidate, image_label),
        },
        "data": {"product": product, "image_asset": candidate, "library_path": str(_library_path(params))},
    }


def update_product_weights(params: dict[str, Any]) -> dict[str, Any]:
    params = params or {}
    matches = find_products(params)
    if not matches:
        return {
            "ok": False,
            "kind": "product_weight_update_error",
            "title": "产品重量更新失败",
            "message": "没有找到要更新重量的产品。请提供工厂名称 + 货号。",
        }
    if len(matches) > 1:
        return {
            "ok": False,
            "kind": "product_weight_update_ambiguous",
            "title": "产品重量更新失败",
            "message": "货号匹配到多个产品，请补充工厂名称后再写入重量。",
            "data": {"matches": matches[:20]},
        }

    updates = _weight_updates_from_params(params)
    if not updates:
        return {
            "ok": False,
            "kind": "product_weight_update_error",
            "title": "产品重量更新失败",
            "message": "请提供单品净重、单品毛重或快递包装重量，例如 350g、0.35kg。",
        }

    product = dict(matches[0])
    manual_specs = dict(_product_manual_specs(product))
    manual_specs.update(updates)
    manual_specs["updated_at"] = _utc_now()
    product["manual_specs"] = manual_specs
    _update_product_in_library(product, params)

    weight_profile = _product_weight_profile(product)
    row = [
        product.get("sku", ""),
        product.get("product_name", ""),
        _single_weight_text(weight_profile, "single_net_weight_g", "single_net_weight_source"),
        _single_weight_text(weight_profile, "single_gross_weight_g", "single_gross_weight_source"),
        _format_grams_with_unit(weight_profile.get("shipping_packaged_weight_g")),
        "已写入产品资料库",
    ]
    return {
        "ok": True,
        "kind": "product_weight_update",
        "title": "产品重量更新结果",
        "columns": ["货号", "品名", "单品净重", "单品毛重", "快递包装重量", "状态"],
        "rows": [row],
        "summary": {
            "工厂": product.get("factory_name"),
            "货号": product.get("sku"),
            "品名": product.get("product_name"),
            "单品净重": row[2],
            "单品毛重": row[3],
            "快递包装重量": row[4],
            "资料库": str(_library_path(params)),
            "状态": "已写入产品资料库，未写入长期记忆",
        },
        "data": {"product": product, "weight_profile": weight_profile, "library_path": str(_library_path(params))},
    }


def update_product_specs(params: dict[str, Any]) -> dict[str, Any]:
    params = params or {}
    matches = find_products(params)
    if not matches:
        return {
            "ok": False,
            "kind": "product_specs_update_error",
            "title": "产品规格更新失败",
            "message": "没有找到要更新规格的产品。请提供工厂名称 + 货号。",
        }
    if len(matches) > 1:
        return {
            "ok": False,
            "kind": "product_specs_update_ambiguous",
            "title": "产品规格更新失败",
            "message": "货号匹配到多个产品，请补充工厂名称后再写入规格。",
            "data": {"matches": matches[:20]},
        }

    updates = _spec_updates_from_params(params)
    if not updates:
        return {
            "ok": False,
            "kind": "product_specs_update_error",
            "title": "产品规格更新失败",
            "message": "请提供产品尺寸、包装尺寸/包装规格或箱规尺寸，例如 28.5*15*3。",
        }

    product = dict(matches[0])
    product.update(updates)
    manual_specs = dict(_product_manual_specs(product))
    manual_specs["specs_updated_at"] = _utc_now()
    product["manual_specs"] = manual_specs
    product["content_hash"] = _content_hash(product)
    _update_product_in_library(product, params)

    row = [
        product.get("sku", ""),
        product.get("product_name", ""),
        _format_cm_with_unit(product.get("product_size_cm")),
        _format_cm_with_unit(product.get("package_size_cm")),
        _format_cm_with_unit(product.get("carton_size_cm")),
        "已写入产品资料库",
    ]
    return {
        "ok": True,
        "kind": "product_specs_update",
        "title": "产品规格更新结果",
        "columns": ["货号", "品名", "产品尺寸", "包装尺寸", "箱规尺寸", "状态"],
        "rows": [row],
        "summary": {
            "工厂": product.get("factory_name"),
            "货号": product.get("sku"),
            "品名": product.get("product_name"),
            "产品尺寸": row[2],
            "包装尺寸": row[3],
            "箱规尺寸": row[4],
            "资料库": str(_library_path(params)),
            "状态": "已写入产品资料库，未写入长期记忆",
        },
        "data": {"product": product, "updated_fields": sorted(updates), "library_path": str(_library_path(params))},
    }


def run(params: dict[str, Any]) -> dict[str, Any]:
    operation = _clean_text((params or {}).get("operation") or (params or {}).get("action") or "query").lower()
    if operation in {"import", "preview", "load"}:
        return import_factory_quotes(params or {})
    if operation in {"status", "library_status", "stats"}:
        return library_status(params or {})
    if operation in {"contact", "query_contact", "factory_contact"}:
        return query_factory_contact(params or {})
    if operation in {"bind_image", "confirm_image_binding", "confirm_image", "attach_image"}:
        return bind_image_asset(params or {})
    if operation in {"update_weight", "set_weight", "bind_weight", "save_weight"}:
        return update_product_weights(params or {})
    if operation in {"update_specs", "set_specs", "update_dimensions", "set_dimensions"}:
        return update_product_specs(params or {})
    if operation in {"image", "image_candidate", "1688", "album"}:
        return create_image_asset_candidate(params or {})
    return query_factory_quote(params or {})
