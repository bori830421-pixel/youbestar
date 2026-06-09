from __future__ import annotations

import base64
import hashlib
import json
import mimetypes
import re
import shutil
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

from core.business_records_db import BusinessRecordsDB, clean_text
from core.http_client import HttpClientError, fetch_bytes, fetch_text
from core.local_runtime import ensure_local_runtime_dirs, local_runtime_dir


DEFAULT_CACHE_LIMIT_BYTES = 500 * 1024 * 1024
RECENT_CAPTURE_LIMIT = 20
SUPPORTED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}
DEFAULT_EXPORT_FIELDS = ["sku_name", "cost_price", "customer_price", "stock", "sku_image_url", "source_url"]
FIELD_LABELS = {
    "index": "序号",
    "sku_name": "SKU名称",
    "sku_id": "SKU ID",
    "spec": "规格",
    "cost_price": "成本价",
    "customer_price": "客户价",
    "price": "价格",
    "stock": "库存",
    "sku_image": "SKU图片",
    "sku_image_url": "图片链接",
    "source_url": "来源链接",
    "matched_record": "匹配资料",
}


class ReferenceProductError(ValueError):
    pass


@dataclass
class JsonCandidate:
    data: Any
    score: int


def run(params: dict[str, Any] | None = None) -> dict[str, Any]:
    params = params or {}
    operation = clean_text(params.get("operation") or params.get("action") or "capture").lower()
    try:
        if operation in {"capture", "read", "fetch", "light_capture"}:
            return capture_reference_product(params)
        if operation in {"match", "candidates", "compare"}:
            return match_reference_product(params)
        if operation in {"confirm_bind", "bind", "write", "save"}:
            return confirm_bind_reference_product(params)
        if operation in {"export_excel", "export", "excel"}:
            return export_reference_product_excel(params)
        if operation in {"cache_status", "status", "info"}:
            return cache_status(params)
        if operation in {"cleanup_cache", "cleanup", "clean"}:
            return cleanup_cache(params)
        raise ReferenceProductError(f"不支持的参考商品操作：{operation}")
    except Exception as exc:
        return {
            "ok": False,
            "kind": "reference_product_error",
            "title": "参考商品处理失败",
            "error": str(exc),
            "message": str(exc),
        }


def capture_reference_product(params: dict[str, Any]) -> dict[str, Any]:
    source_url = _source_url(params)
    html = _source_html(params)
    payload = params.get("data") or params.get("payload") or params.get("structured_data")
    if not isinstance(payload, (dict, list)):
        payload = None

    if html is None and payload is None:
        _validate_reference_url(source_url)
        html = fetch_text(source_url, timeout=int(params.get("timeout") or 15))

    parsed = normalize_reference_product(
        payload if payload is not None else html or "",
        source_url=source_url,
        title_hint=clean_text(params.get("title") or params.get("categoryKeyword")),
    )
    if not parsed["skus"]:
        raise ReferenceProductError("没有读取到 SKU 信息；请换一个 1688 商品详情链接，或先打开页面后再试。")

    capture_id = _capture_id(source_url, parsed["title"], parsed["skus"])
    record = {
        "capture_id": capture_id,
        "offer_id": _offer_id(source_url, parsed),
        "source_url": source_url,
        "title": parsed["title"],
        "attributes": parsed.get("attributes", {}),
        "skus": parsed["skus"],
        "resource_hints": parsed.get("resource_hints", {}),
        "captured_at": utc_now(),
    }
    capture_path = captures_dir() / f"{capture_id}.json"
    _write_json(capture_path, record)
    _update_state({"latest_capture_id": capture_id}, recent_capture_id=capture_id)

    skus = record["skus"]
    rows = [
        [
            index + 1,
            sku.get("sku_name") or "-",
            sku.get("cost_price_text") or sku.get("price") or "-",
            sku.get("stock", "-"),
            _image_markdown(sku.get("image_url") or sku.get("sku_image_url")),
        ]
        for index, sku in enumerate(skus[:20])
    ]
    return {
        "ok": True,
        "kind": "reference_product_capture",
        "title": "参考商品已读取",
        "columns": ["序号", "SKU名称", "价格", "库存", "图片"],
        "rows": rows,
        "summary": {
            "商品标题": record["title"] or "-",
            "SKU数量": len(skus),
            "图片URL数量": sum(1 for sku in skus if sku.get("sku_image_url")),
            "临时采集ID": capture_id,
        },
        "capture": record,
        "capture_id": capture_id,
        "capture_path": str(capture_path),
        "data": record,
        "next_actions": ["生成候选", "导出内部 Excel", "导出客户报价 Excel"],
    }


def match_reference_product(params: dict[str, Any]) -> dict[str, Any]:
    capture = _load_capture_from_params(params)
    query = clean_text(params.get("query") or params.get("keyword") or capture.get("title"))
    limit = int(params.get("limit") or 100)
    database = _db(params)
    records = database.query_records(record_type="product", query="", limit=limit)
    if query and not records:
        records = database.query_records(record_type="product", query=query, limit=limit)

    matches = []
    for sku in capture.get("skus", []):
        candidates = sorted(
            (
                _score_record_candidate(sku, record, capture)
                for record in records
            ),
            key=lambda item: item["confidence"],
            reverse=True,
        )
        candidates = [candidate for candidate in candidates if candidate["confidence"] > 0][:5]
        best = candidates[0] if candidates else None
        matches.append(
            {
                "source_sku_id": sku.get("source_sku_id") or sku.get("sku_id") or "",
                "sku": sku.get("sku") or sku.get("sku_id") or "",
                "sku_name": sku.get("sku_name") or "",
                "sku_id": sku.get("sku_id") or "",
                "image_url": sku.get("image_url") or sku.get("sku_image_url") or "",
                "sku_image_url": sku.get("sku_image_url") or sku.get("image_url") or "",
                "price": sku.get("price") or sku.get("cost_price_text") or "",
                "cost_price": sku.get("cost_price", ""),
                "stock": sku.get("stock", ""),
                "best_candidate": best,
                "candidates": candidates,
                "status": "matched" if best and best["confidence"] >= 0.85 else "needs_confirmation" if best else "new_candidate",
            }
        )

    match_id = _match_id(capture["capture_id"], matches)
    match_record = {
        "match_id": match_id,
        "capture_id": capture["capture_id"],
        "source_url": capture.get("source_url", ""),
        "title": capture.get("title", ""),
        "matches": matches,
        "created_at": utc_now(),
    }
    match_path = matches_dir() / f"{match_id}.json"
    _write_json(match_path, match_record)
    _update_state({"latest_match_id": match_id, "latest_capture_id": capture["capture_id"]})

    rows = []
    for item in matches[:30]:
        candidate = item.get("best_candidate") or {}
        rows.append(
            [
                item.get("sku_name") or "-",
                item.get("price") or "-",
                _image_markdown(item.get("image_url") or item.get("sku_image_url")),
                candidate.get("title") or "待新建/待确认",
                f"{float(candidate.get('confidence') or candidate.get('score') or 0):.0%}",
                item.get("status") or "-",
            ]
        )

    result = {
        "ok": True,
        "kind": "reference_product_match",
        "title": "参考商品候选已生成",
        "columns": ["SKU名称", "价格", "图片", "候选资料", "置信度", "状态"],
        "rows": rows,
        "summary": {
            "临时采集ID": capture["capture_id"],
            "候选批次": match_id,
            "SKU数量": len(matches),
            "高置信匹配": sum(1 for item in matches if (item.get("best_candidate") or {}).get("confidence", 0) >= 0.85),
            "待确认": sum(1 for item in matches if item.get("status") != "matched"),
        },
        "match": match_record,
        "match_id": match_id,
        "match_path": str(match_path),
        "matches": matches,
        "data": match_record,
    }
    from core.ui_interactions import build_chat_interaction

    interaction = build_chat_interaction("official.reference_product", result)
    if interaction:
        result["interactions"] = [interaction]
    return result


def confirm_bind_reference_product(params: dict[str, Any]) -> dict[str, Any]:
    if params.get("confirmed") is not True and params.get("confirm") is not True:
        return {
            "ok": False,
            "kind": "reference_product_bind_confirmation_required",
            "title": "参考商品关联需要确认",
            "message": "写入资料库前需要明确确认 confirmed=true。",
            "summary": {"状态": "待确认"},
        }

    direct_binding = _direct_binding_from_params(params)
    if direct_binding:
        match = {"match_id": "", "matches": []}
        capture = {
            "capture_id": "",
            "source_url": clean_text(params.get("source_url") or params.get("reference_product_url")),
            "title": clean_text(params.get("title")),
            "skus": [direct_binding],
        }
    else:
        match = _load_match_from_params(params)
        capture = _load_capture(match["capture_id"])
    database = _db(params)
    bindings = [direct_binding] if direct_binding else _resolve_bindings(params, match)
    if not bindings:
        raise ReferenceProductError("没有可写入的候选。请先生成候选，或传入 bindings。")

    saved_records = []
    for binding in bindings:
        sku = _find_sku(capture, binding.get("sku_name"), binding.get("sku_id") or binding.get("source_sku_id"))
        record_id = clean_text(binding.get("record_id") or binding.get("id"))
        candidate = binding.get("candidate") if isinstance(binding.get("candidate"), dict) else {}
        if not record_id:
            record_id = clean_text(candidate.get("record_id") or candidate.get("id"))
        if not record_id and not params.get("create_missing"):
            continue

        image_url = sku.get("image_url") or sku.get("sku_image_url") or binding.get("image_url") or ""
        source_sku_id = sku.get("source_sku_id") or sku.get("sku_id") or binding.get("source_sku_id") or binding.get("sku_id") or ""
        sku_code = sku.get("sku") or sku.get("sku_id") or binding.get("sku") or source_sku_id or candidate.get("business_key") or ""
        sku_name = sku.get("sku_name") or binding.get("sku_name") or candidate.get("title") or capture.get("title") or ""
        fields = {
            "sku": sku_code,
            "name": sku_name,
            "reference_sku_name": sku_name,
            "reference_sku_id": source_sku_id,
            "reference_price": sku.get("cost_price") if sku.get("cost_price") not in (None, "") else sku.get("price") or "",
            "reference_stock": sku.get("stock", ""),
            "sku_image_url": image_url,
            "reference_image_url": image_url,
            "reference_product_url": capture.get("source_url") or "",
            "reference_capture_id": capture.get("capture_id") or "",
        }
        fields = {key: value for key, value in fields.items() if value not in (None, "")}
        saved = database.upsert_record(
            record_type="product",
            fields=fields,
            title=fields.get("name") or None,
            business_key=clean_text(binding.get("business_key") or candidate.get("business_key")) or None,
            source="official.reference_product",
            actor=clean_text(params.get("actor") or "operator"),
            source_ip=clean_text(params.get("source_ip") or ""),
            record_id=record_id or None,
        )
        saved_records.append(saved)

    rows = [
        [
            record.get("business_key") or "-",
            record.get("title") or "-",
            (record.get("fields") or {}).get("sku_image_url") or "-",
            record.get("action") or "-",
        ]
        for record in saved_records
    ]
    return {
        "ok": True,
        "kind": "reference_product_bind",
        "title": "参考商品图片关联已确认",
        "columns": ["业务键", "产品", "图片链接", "动作"],
        "rows": rows,
        "summary": {
            "写入记录数": len(saved_records),
            "候选批次": match.get("match_id") or "-",
            "临时采集ID": capture.get("capture_id") or "-",
        },
        "records": saved_records,
        "data": {"records": saved_records, "match_id": match.get("match_id"), "capture_id": capture.get("capture_id")},
    }


def export_reference_product_excel(params: dict[str, Any]) -> dict[str, Any]:
    capture = _load_capture_from_params(params)
    match = _load_optional_match(params)
    export_mode = clean_text(params.get("mode") or params.get("export_mode") or "internal").lower()
    margin_percent = _margin_percent(params)
    decimal_places = max(0, min(int(params.get("decimal_places") or 2), 6))
    embed_images = bool(params.get("embed_images") or params.get("include_images"))
    fields = _export_fields(params, export_mode, embed_images)
    output_path = _resolve_output_path(params, capture)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = _export_rows(capture, match, fields, export_mode, margin_percent, decimal_places)
    downloaded_images = {}
    if embed_images:
        downloaded_images = _download_export_images(capture, rows)
    _write_export_workbook(output_path, fields, rows, downloaded_images, embed_images)
    _update_state({"latest_export_path": str(output_path), "latest_capture_id": capture["capture_id"]})

    return {
        "ok": True,
        "kind": "reference_product_export",
        "title": "参考商品 Excel 已导出",
        "columns": ["文件", "行数", "模式", "加点", "图片"],
        "rows": [[str(output_path), len(rows), export_mode, f"{margin_percent}%", "已尝试嵌入" if embed_images else "仅链接"]],
        "summary": {
            "导出文件": str(output_path),
            "SKU行数": len(rows),
            "导出模式": "客户报价" if export_mode == "customer" else "内部查看",
            "加点": f"{margin_percent}%",
            "小数位": decimal_places,
            "图片缓存数": len(downloaded_images),
            "客户报价": _first_customer_price_text(rows),
        },
        "output_path": str(output_path),
        "row_count": len(rows),
        "image_downloaded_count": len(downloaded_images),
        "data": {"output_path": str(output_path), "rows": rows, "items": rows, "fields": fields},
    }


def cache_status(params: dict[str, Any] | None = None) -> dict[str, Any]:
    params = params or {}
    root = cache_root()
    root.mkdir(parents=True, exist_ok=True)
    total_size, file_count = _cache_size(root)
    max_bytes = int(params.get("max_bytes") or DEFAULT_CACHE_LIMIT_BYTES)
    return {
        "ok": True,
        "kind": "reference_product_cache_status",
        "title": "参考商品缓存状态",
        "columns": ["缓存目录", "文件数", "占用", "上限"],
        "rows": [[str(root), file_count, _format_bytes(total_size), _format_bytes(max_bytes)]],
        "summary": {
            "缓存目录": str(root),
            "文件数": file_count,
            "占用": _format_bytes(total_size),
            "上限": _format_bytes(max_bytes),
        },
        "data": {"cache_dir": str(root), "size_bytes": total_size, "file_count": file_count, "max_bytes": max_bytes},
    }


def cleanup_cache(params: dict[str, Any] | None = None) -> dict[str, Any]:
    params = params or {}
    root = cache_root()
    root.mkdir(parents=True, exist_ok=True)
    max_bytes = int(params.get("max_bytes") or DEFAULT_CACHE_LIMIT_BYTES)
    total_size, _ = _cache_size(root)
    removed = []
    if total_size > max_bytes:
        files = sorted(
            (item for item in root.rglob("*") if item.is_file()),
            key=lambda item: item.stat().st_mtime,
        )
        for file_path in files:
            if total_size <= max_bytes:
                break
            size = file_path.stat().st_size
            file_path.unlink()
            removed.append({"path": str(file_path), "size": size})
            total_size -= size
    _remove_empty_dirs(root)
    final_size, final_count = _cache_size(root)
    return {
        "ok": True,
        "kind": "reference_product_cache_cleanup",
        "title": "参考商品缓存已清理",
        "columns": ["删除文件数", "释放空间", "当前占用"],
        "rows": [[len(removed), _format_bytes(sum(item["size"] for item in removed)), _format_bytes(final_size)]],
        "summary": {
            "删除文件数": len(removed),
            "释放空间": _format_bytes(sum(item["size"] for item in removed)),
            "当前占用": _format_bytes(final_size),
            "当前文件数": final_count,
        },
        "removed": removed,
        "data": {"removed": removed, "size_bytes": final_size, "file_count": final_count},
    }


def normalize_reference_product(raw: Any, source_url: str = "", title_hint: str = "") -> dict[str, Any]:
    if isinstance(raw, str):
        html = raw
        candidates = _json_candidates_from_html(html)
        structured = _best_structured_candidate(candidates)
        title = _extract_title(structured, html, title_hint)
        attributes = _extract_attributes(structured)
        resource_hints = _extract_resource_hints(structured, html, source_url)
    else:
        html = ""
        structured = raw
        title = _extract_title(structured, "", title_hint)
        attributes = _extract_attributes(structured)
        resource_hints = _extract_resource_hints(structured, "", source_url)

    sku_dimensions = _extract_sku_dimensions(structured)
    sku_rows = _extract_sku_rows(structured, sku_dimensions)
    if not sku_rows and sku_dimensions:
        sku_rows = _sku_rows_from_dimensions(sku_dimensions)
    if not sku_rows and isinstance(structured, dict):
        direct_skus = structured.get("skus") or structured.get("skuRows") or structured.get("sku_rows")
        sku_rows = _normalize_direct_skus(direct_skus, sku_dimensions)

    image_urls = resource_hints.get("skuImages") or resource_hints.get("mainImages") or []
    for index, sku in enumerate(sku_rows):
        if not sku.get("sku_image_url") and index < len(image_urls):
            sku["sku_image_url"] = image_urls[index]
            sku["image_url"] = image_urls[index]
        if title and sku.get("props") and title not in clean_text(sku.get("sku_name")):
            sku["sku_name"] = clean_text(f"{title} {sku.get('sku_name')}")

    return {
        "offer_id": clean_text(_find_first_key(structured, "offerId") or _find_first_key(structured, "offer_id")),
        "title": title,
        "attributes": attributes,
        "skus": sku_rows,
        "resource_hints": resource_hints,
    }


def _source_url(params: dict[str, Any]) -> str:
    return clean_text(
        params.get("source_url")
        or params.get("reference_url")
        or params.get("referenceOfferUrl")
        or params.get("url")
        or params.get("link")
    )


def _source_html(params: dict[str, Any]) -> str | None:
    value = params.get("html") or params.get("source_html") or params.get("page_html")
    if isinstance(value, str) and value.strip():
        return value
    path = clean_text(params.get("html_path") or params.get("source_html_path"))
    if path:
        return Path(path).read_text(encoding="utf-8")
    return None


def _validate_reference_url(url: str) -> None:
    if not url:
        raise ReferenceProductError("请提供 1688 商品详情链接。")
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ReferenceProductError("参考商品链接不是有效的网址。")
    if "1688.com" not in parsed.netloc:
        raise ReferenceProductError("目前参考商品采集只支持 1688 商品链接。")


def cache_root() -> Path:
    ensure_local_runtime_dirs()
    return local_runtime_dir() / "cache" / "reference_products"


def captures_dir() -> Path:
    path = cache_root() / "captures"
    path.mkdir(parents=True, exist_ok=True)
    return path


def matches_dir() -> Path:
    path = cache_root() / "matches"
    path.mkdir(parents=True, exist_ok=True)
    return path


def images_dir() -> Path:
    path = cache_root() / "images"
    path.mkdir(parents=True, exist_ok=True)
    return path


def exports_dir() -> Path:
    path = local_runtime_dir() / "exports" / "reference_products"
    path.mkdir(parents=True, exist_ok=True)
    return path


def state_file() -> Path:
    return cache_root() / "state.json"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _timestamp_for_file() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _load_state() -> dict[str, Any]:
    try:
        data = _read_json(state_file())
    except (FileNotFoundError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def _update_state(updates: dict[str, Any], recent_capture_id: str | None = None) -> None:
    state = _load_state()
    state.update(updates)
    if recent_capture_id:
        recent = [item for item in state.get("recent_capture_ids", []) if item != recent_capture_id]
        state["recent_capture_ids"] = [recent_capture_id, *recent][:RECENT_CAPTURE_LIMIT]
    state["updated_at"] = utc_now()
    _write_json(state_file(), state)


def _capture_id(source_url: str, title: str, skus: list[dict[str, Any]]) -> str:
    digest = hashlib.sha1(json.dumps([source_url, title, skus[:5]], ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()[:10]
    return f"ref-{datetime.now().strftime('%Y%m%d%H%M%S')}-{digest}"


def _offer_id(source_url: str, parsed: dict[str, Any]) -> str:
    direct = clean_text(parsed.get("offer_id") or parsed.get("offerId"))
    if direct:
        return direct
    match = re.search(r"/offer/(\d+)\.html", source_url)
    return match.group(1) if match else ""


def _match_id(capture_id: str, matches: list[dict[str, Any]]) -> str:
    digest = hashlib.sha1(json.dumps(matches[:10], ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()[:8]
    return f"match-{capture_id}-{digest}"


def _load_capture_from_params(params: dict[str, Any]) -> dict[str, Any]:
    capture = params.get("capture") or params.get("reference_product") or params.get("data")
    if isinstance(capture, dict) and capture.get("skus"):
        return capture
    direct_skus = params.get("skus") or params.get("items")
    if isinstance(direct_skus, list) and direct_skus:
        normalized = normalize_reference_product(
            {
                "title": params.get("title") or params.get("product_title") or "",
                "skus": direct_skus,
            },
            source_url=_source_url(params),
        )
        capture_id = clean_text(params.get("capture_id")) or _capture_id(_source_url(params), normalized["title"], normalized["skus"])
        return {
            "capture_id": capture_id,
            "offer_id": _offer_id(_source_url(params), normalized),
            "source_url": _source_url(params),
            "title": normalized["title"],
            "attributes": normalized.get("attributes", {}),
            "skus": normalized["skus"],
            "resource_hints": normalized.get("resource_hints", {}),
            "captured_at": utc_now(),
        }
    capture_id = clean_text(params.get("capture_id") or params.get("reference_capture_id"))
    if not capture_id:
        capture_id = clean_text(_load_state().get("latest_capture_id"))
    if not capture_id:
        raise ReferenceProductError("没有可复用的参考商品采集结果，请先读取参考商品。")
    return _load_capture(capture_id)


def _load_capture(capture_id: str) -> dict[str, Any]:
    path = captures_dir() / f"{capture_id}.json"
    if not path.exists():
        raise ReferenceProductError(f"找不到参考商品采集结果：{capture_id}")
    data = _read_json(path)
    if not isinstance(data, dict) or not isinstance(data.get("skus"), list):
        raise ReferenceProductError(f"参考商品采集结果无效：{capture_id}")
    return data


def _load_match_from_params(params: dict[str, Any]) -> dict[str, Any]:
    match = params.get("match")
    if isinstance(match, dict) and match.get("matches"):
        return match
    match_id = clean_text(params.get("match_id") or params.get("candidate_id"))
    if not match_id:
        match_id = clean_text(_load_state().get("latest_match_id"))
    if not match_id:
        raise ReferenceProductError("没有可复用的候选结果，请先生成候选。")
    path = matches_dir() / f"{match_id}.json"
    if not path.exists():
        raise ReferenceProductError(f"找不到候选结果：{match_id}")
    data = _read_json(path)
    if not isinstance(data, dict) or not isinstance(data.get("matches"), list):
        raise ReferenceProductError(f"候选结果无效：{match_id}")
    return data


def _load_optional_match(params: dict[str, Any]) -> dict[str, Any] | None:
    try:
        return _load_match_from_params(params)
    except ReferenceProductError:
        return None


def _db(params: dict[str, Any]) -> BusinessRecordsDB:
    return BusinessRecordsDB(
        db_path=params.get("database_path"),
        audit_log_path=params.get("audit_log_path"),
        schema_path=params.get("schema_path"),
    )


def _json_candidates_from_html(html: str) -> list[JsonCandidate]:
    candidates: list[JsonCandidate] = []
    for match in re.finditer(r"<script[^>]*type=[\"']application/(?:ld\+)?json[\"'][^>]*>(.*?)</script>", html, re.I | re.S):
        data = _loads_jsonish(_html_unescape(match.group(1)))
        if data is not None:
            candidates.append(JsonCandidate(data=data, score=_structured_score(data) + 5))

    markers = ("window.context", "__INIT_DATA", "__INITIAL_STATE", "offerData", "dataJson", "skuProps", "skuMap")
    for marker in markers:
        index = html.find(marker)
        while index != -1:
            brace_index = html.find("{", index)
            if brace_index == -1:
                break
            text = _extract_balanced_json(html, brace_index)
            data = _loads_jsonish(text)
            if data is not None:
                candidates.append(JsonCandidate(data=data, score=_structured_score(data)))
            index = html.find(marker, index + len(marker))

    for match in re.finditer(r"(\{[^{}]*(?:skuProps|skuMap|offerTitle|skuMapOriginal)[\s\S]{0,20000}\})", html, re.I):
        data = _loads_jsonish(match.group(1))
        if data is not None:
            candidates.append(JsonCandidate(data=data, score=_structured_score(data)))
    return candidates


def _best_structured_candidate(candidates: list[JsonCandidate]) -> Any:
    if not candidates:
        return {}
    return max(candidates, key=lambda item: item.score).data


def _structured_score(data: Any) -> int:
    text = json.dumps(data, ensure_ascii=False, default=str)[:200000]
    score = 0
    for pattern, value in (
        ("skuProps", 8),
        ("skuMap", 8),
        ("skuMapOriginal", 10),
        ("offerTitle", 4),
        ("imageUrl", 3),
        ("price", 2),
        ("stock", 2),
    ):
        if pattern in text:
            score += value
    return score


def _extract_balanced_json(text: str, start: int) -> str:
    depth = 0
    in_string = False
    quote = ""
    escaped = False
    for index in range(start, min(len(text), start + 500000)):
        char = text[index]
        if in_string:
            if escaped:
                escaped = False
            elif char == "\\":
                escaped = True
            elif char == quote:
                in_string = False
            continue
        if char in {"'", '"'}:
            in_string = True
            quote = char
        elif char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return text[start : index + 1]
    return text[start : start + 500000]


def _loads_jsonish(text: str) -> Any:
    clean = text.strip()
    if not clean:
        return None
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        pass
    quoted = re.sub(r"(?<![\"'])\b([A-Za-z_][A-Za-z0-9_]*)\b\s*:", r'"\1":', clean)
    quoted = quoted.replace("'", '"')
    quoted = re.sub(r",\s*([}\]])", r"\1", quoted)
    try:
        return json.loads(quoted)
    except json.JSONDecodeError:
        return None


def _html_unescape(value: str) -> str:
    return (
        value.replace("&quot;", '"')
        .replace("&#34;", '"')
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
    )


def _extract_title(data: Any, html: str, title_hint: str = "") -> str:
    if title_hint:
        return title_hint
    for key in ("offerTitle", "subject", "title", "name", "productTitle"):
        value = _find_first_key(data, key)
        if isinstance(value, str) and clean_text(value):
            return clean_text(value)
        if isinstance(value, dict):
            nested = value.get("fields", {}).get("title") if isinstance(value.get("fields"), dict) else value.get("title")
            if clean_text(nested):
                return clean_text(nested)
    match = re.search(r"<title[^>]*>(.*?)</title>", html, re.I | re.S)
    if match:
        return clean_text(re.sub(r"<[^>]+>", "", _html_unescape(match.group(1))).replace("- 阿里巴巴", ""))
    return ""


def _extract_attributes(data: Any) -> dict[str, str]:
    attributes: dict[str, str] = {}
    raw = _find_first_key(data, "attributes") or _find_first_key(data, "productAttributes")
    if isinstance(raw, dict):
        for key, value in raw.items():
            if clean_text(key) and clean_text(value):
                attributes[clean_text(key)] = clean_text(value)
    elif isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            name = clean_text(item.get("name") or item.get("key") or item.get("attributeName") or item.get("title"))
            value = clean_text(item.get("value") or item.get("valueName") or item.get("text"))
            if name and value:
                attributes[name] = value
    return attributes


def _extract_resource_hints(data: Any, html: str, source_url: str) -> dict[str, list[str]]:
    main_images = []
    sku_images = []
    videos = []
    for url in _find_image_urls(data):
        if "video" in url.lower():
            videos.append(_normalize_url(url, source_url))
        elif "sku" in url.lower() or "_sum" in url.lower():
            sku_images.append(_normalize_image_url(url, source_url))
        else:
            main_images.append(_normalize_image_url(url, source_url))
    if html:
        for url in re.findall(r"https?://[^\"'\s<>]+?\.(?:jpg|jpeg|png|webp|gif)(?:\?[^\"'\s<>]*)?", html, re.I):
            main_images.append(_normalize_image_url(url, source_url))
        for url in re.findall(r"(?:src|data-src|imageUrl|imgUrl)=[\"']([^\"']+\.(?:jpg|jpeg|png|webp|gif)(?:\?[^\"']*)?)[\"']", html, re.I):
            main_images.append(_normalize_image_url(url, source_url))
    return {
        "mainImages": _unique_urls(main_images),
        "skuImages": _unique_urls(sku_images),
        "videos": _unique_urls(videos),
        "detailImages": [],
    }


def _extract_sku_dimensions(data: Any) -> list[dict[str, Any]]:
    raw = _find_first_key(data, "skuProps") or _find_first_key(data, "sku_props")
    if raw is None:
        raw = _find_path(data, ("skuModel", "skuProps"))
    groups = raw if isinstance(raw, list) else list(raw.values()) if isinstance(raw, dict) else []
    dimensions = []
    for group_index, group in enumerate(groups, start=1):
        if not isinstance(group, dict):
            continue
        values = group.get("value") or group.get("values") or group.get("skuValues") or group.get("propValues") or group.get("children") or []
        value_rows = []
        for value_index, item in enumerate(values if isinstance(values, list) else list(values.values()) if isinstance(values, dict) else [], start=1):
            if not isinstance(item, dict):
                continue
            name = clean_text(item.get("name") or item.get("value") or item.get("valueName") or item.get("text") or item.get("title"))
            if not name:
                continue
            value_rows.append(
                {
                    "index": value_index,
                    "value_id": clean_text(item.get("valueId") or item.get("vid") or item.get("id") or item.get("skuValueId")),
                    "name": name,
                    "image_url": _normalize_image_url(
                        item.get("imageUrl") or item.get("image") or item.get("imgUrl") or item.get("skuImageUrl") or item.get("bigImageUrl") or "",
                        "",
                    ),
                }
            )
        name = clean_text(group.get("prop") or group.get("propName") or group.get("name") or group.get("skuPropName") or group.get("attributeName") or f"规格{group_index}")
        if name and value_rows:
            dimensions.append({"index": group_index, "name": name, "values": value_rows})
    return dimensions


def _extract_sku_rows(data: Any, dimensions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    raw = (
        _find_first_key(data, "skuMapOriginal")
        or _find_first_key(data, "skuMapOrigin")
        or _find_first_key(data, "skuMap")
        or _find_path(data, ("mainPrice", "fields", "finalPriceModel", "tradeWithoutPromotion", "skuMapOriginal"))
        or _find_path(data, ("Root", "fields", "dataJson", "skuModel", "skuMap"))
    )
    return _normalize_direct_skus(raw, dimensions)


def _normalize_direct_skus(raw: Any, dimensions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    raw_items = []
    if isinstance(raw, list):
        raw_items = [(None, item) for item in raw]
    elif isinstance(raw, dict):
        raw_items = list(raw.items())
    rows = []
    for index, (raw_key, item) in enumerate(raw_items, start=1):
        if not isinstance(item, dict):
            continue
        name = clean_text(item.get("specAttrs") or item.get("specAttr") or item.get("name") or item.get("skuName") or item.get("sku_name") or raw_key)
        spec_values = _match_spec_values(name, dimensions)
        image_url = clean_text(item.get("imageUrl") or item.get("image") or item.get("imgUrl") or item.get("skuImageUrl"))
        if not image_url:
            image_url = next((value.get("image_url") for value in spec_values if value.get("image_url")), "")
        if not name and spec_values:
            name = " / ".join(value["value"] for value in spec_values)
        if not name and not clean_text(item.get("skuId")):
            continue
        props = {value.get("group", ""): value.get("value", "") for value in spec_values if value.get("group") and value.get("value")}
        sku_code = clean_text(item.get("skuCode") or item.get("sku") or item.get("code") or item.get("outerId"))
        source_sku_id = clean_text(item.get("source_sku_id") or item.get("skuId") or item.get("skuID") or item.get("id"))
        price_value = item.get("discountPrice") or item.get("price") or item.get("salePrice") or item.get("cost_price") or item.get("costPrice")
        cost_price = _decimal_or_blank(price_value)
        cost_price_value = float(cost_price) if isinstance(cost_price, Decimal) else ""
        display_name = name
        if props:
            base_title = ""
            display_parts = [value for value in props.values() if value]
            display_name = " ".join(part for part in [base_title, *display_parts] if part) or name
        rows.append(
            {
                "index": index,
                "source_sku_id": source_sku_id,
                "sku_id": source_sku_id,
                "sku": sku_code or source_sku_id,
                "sku_name": display_name,
                "spec": " / ".join(value["value"] for value in spec_values) or name,
                "props": props,
                "spec_values": spec_values,
                "price": clean_text(price_value),
                "cost_price": cost_price_value,
                "cost_price_text": clean_text(price_value),
                "stock": item.get("canBookCount", item.get("stock", item.get("amount", ""))),
                "image_url": _normalize_image_url(item.get("image_url") or image_url, ""),
                "sku_image_url": _normalize_image_url(item.get("sku_image_url") or item.get("image_url") or image_url, ""),
            }
        )
    return rows


def _sku_rows_from_dimensions(dimensions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not dimensions:
        return []
    rows = []
    first = dimensions[0]
    for index, value in enumerate(first["values"], start=1):
        rows.append(
            {
                "index": index,
                "sku_id": value.get("value_id") or "",
                "sku_name": value.get("name") or "",
                "spec": value.get("name") or "",
                "spec_values": [{"group": first.get("name"), "value": value.get("name"), "image_url": value.get("image_url")}],
                "price": "",
                "stock": "",
                "sku_image_url": value.get("image_url") or "",
            }
        )
    return rows


def _match_spec_values(sku_name: str, dimensions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    text = clean_text(sku_name).replace("&gt;", ">")
    matches = []
    for group in dimensions:
        candidates = []
        for value in group.get("values", []):
            value_name = clean_text(value.get("name"))
            if not value_name:
                continue
            score = 0
            if value_name and value_name in text:
                score = len(value_name)
            elif text and text in value_name:
                score = len(text)
            if score:
                candidates.append((score, value))
        if candidates:
            _, best = sorted(candidates, key=lambda item: item[0], reverse=True)[0]
            matches.append(
                {
                    "group": group.get("name") or "",
                    "value": best.get("name") or "",
                    "value_id": best.get("value_id") or "",
                    "image_url": best.get("image_url") or "",
                }
            )
    return matches


def _find_first_key(value: Any, target_key: str) -> Any:
    if isinstance(value, dict):
        if target_key in value:
            return value[target_key]
        for child in value.values():
            found = _find_first_key(child, target_key)
            if found is not None:
                return found
    elif isinstance(value, list):
        for child in value:
            found = _find_first_key(child, target_key)
            if found is not None:
                return found
    return None


def _find_path(value: Any, path: tuple[str, ...]) -> Any:
    current = value
    for part in path:
        if not isinstance(current, dict) or part not in current:
            return None
        current = current[part]
    return current


def _find_image_urls(value: Any) -> list[str]:
    urls: list[str] = []
    if isinstance(value, dict):
        for key, child in value.items():
            if isinstance(child, str) and re.search(r"\.(?:jpg|jpeg|png|webp|gif)(?:$|\?)", child, re.I):
                if re.search(r"(image|img|url|src|video)", str(key), re.I):
                    urls.append(child)
            else:
                urls.extend(_find_image_urls(child))
    elif isinstance(value, list):
        for child in value:
            urls.extend(_find_image_urls(child))
    return urls


def _normalize_url(url: Any, base_url: str = "") -> str:
    text = clean_text(url)
    if not text:
        return ""
    if text.startswith("//"):
        text = f"https:{text}"
    if base_url and not re.match(r"^https?://", text):
        text = urljoin(base_url, text)
    return text


def _normalize_image_url(url: Any, base_url: str = "") -> str:
    text = _normalize_url(url, base_url)
    if not text:
        return ""
    text = text.replace("\\/", "/")
    return text


def _unique_urls(urls: list[str]) -> list[str]:
    result = []
    seen = set()
    for url in urls:
        clean = clean_text(url)
        if not clean or clean in seen:
            continue
        seen.add(clean)
        result.append(clean)
    return result


def _score_record_candidate(sku: dict[str, Any], record: dict[str, Any], capture: dict[str, Any]) -> dict[str, Any]:
    fields = record.get("fields") if isinstance(record.get("fields"), dict) else {}
    record_values = [
        record.get("business_key"),
        record.get("title"),
        fields.get("sku"),
        fields.get("product_code"),
        fields.get("product_id"),
        fields.get("name"),
        fields.get("product_name"),
        fields.get("spec"),
    ]
    sku_keys = [sku.get("sku"), sku.get("sku_id"), sku.get("source_sku_id"), sku.get("sku_name"), sku.get("spec")]
    score = 0.0
    reason = ""
    for left in sku_keys:
        for right in record_values:
            left_norm = _normalize_match_text(left)
            right_norm = _normalize_match_text(right)
            if not left_norm or not right_norm:
                continue
            if left_norm == right_norm:
                score = max(score, 0.95)
                reason = "SKU/名称精确匹配"
            elif left_norm in right_norm or right_norm in left_norm:
                score = max(score, 0.72)
                reason = "SKU/名称包含匹配"
            else:
                ratio = SequenceMatcher(None, left_norm, right_norm).ratio()
                if ratio >= 0.72:
                    score = max(score, ratio * 0.75)
                    reason = "名称相似"
    title_norm = _normalize_match_text(capture.get("title"))
    record_title_norm = _normalize_match_text(record.get("title") or fields.get("name") or fields.get("product_name"))
    if title_norm and record_title_norm and (title_norm in record_title_norm or record_title_norm in title_norm):
        score = max(score, 0.45)
        reason = reason or "商品标题相近"
    return {
        "record_id": record.get("id") or record.get("record_id") or "",
        "id": record.get("id") or record.get("record_id") or "",
        "business_key": record.get("business_key") or "",
        "title": record.get("title") or fields.get("name") or fields.get("product_name") or "",
        "confidence": round(score, 4),
        "score": round(score, 4),
        "reason": reason or "未匹配",
        "fields": fields,
    }


def _normalize_match_text(value: Any) -> str:
    return re.sub(r"[\s【】\[\]（）(),，:：;；/\\|_-]+", "", clean_text(value).lower())


def _resolve_bindings(params: dict[str, Any], match: dict[str, Any]) -> list[dict[str, Any]]:
    bindings = params.get("bindings") or params.get("selected")
    if isinstance(bindings, list) and bindings:
        return [item for item in bindings if isinstance(item, dict)]
    candidate = params.get("candidate") if isinstance(params.get("candidate"), dict) else {}
    best_candidate = candidate.get("best_candidate") if isinstance(candidate.get("best_candidate"), dict) else candidate.get("candidate")
    best_candidate = best_candidate if isinstance(best_candidate, dict) else candidate
    if candidate and clean_text(params.get("record_id") or best_candidate.get("record_id") or best_candidate.get("id")):
        return [
            {
                "sku_name": clean_text(params.get("sku_name") or candidate.get("sku_name")),
                "sku_id": clean_text(params.get("source_sku_id") or candidate.get("source_sku_id") or candidate.get("sku_id")),
                "source_sku_id": clean_text(params.get("source_sku_id") or candidate.get("source_sku_id") or candidate.get("sku_id")),
                "record_id": clean_text(params.get("record_id") or best_candidate.get("record_id") or best_candidate.get("id")),
                "business_key": clean_text(params.get("business_key") or best_candidate.get("business_key")),
                "candidate": best_candidate,
            }
        ]
    threshold = float(params.get("min_confidence") or params.get("auto_bind_min_confidence") or 0.85)
    resolved = []
    for item in match.get("matches", []):
        candidate = item.get("best_candidate") if isinstance(item, dict) else None
        if not isinstance(candidate, dict) or float(candidate.get("confidence") or 0) < threshold:
            continue
        resolved.append(
            {
                "sku_name": item.get("sku_name") or "",
                "sku_id": item.get("sku_id") or "",
                "record_id": candidate.get("record_id") or "",
                "business_key": candidate.get("business_key") or "",
                "candidate": candidate,
            }
        )
    return resolved


def _find_sku(capture: dict[str, Any], sku_name: Any, sku_id: Any) -> dict[str, Any]:
    clean_name = clean_text(sku_name)
    clean_id = clean_text(sku_id)
    for sku in capture.get("skus", []):
        if clean_id and clean_id in {clean_text(sku.get("sku_id")), clean_text(sku.get("source_sku_id"))}:
            return sku
        if clean_name and clean_text(sku.get("sku_name")) == clean_name:
            return sku
    return {}


def _direct_binding_from_params(params: dict[str, Any]) -> dict[str, Any] | None:
    record_id = clean_text(params.get("record_id") or params.get("id"))
    sku = clean_text(params.get("sku"))
    image_url = clean_text(params.get("image_url") or params.get("sku_image_url") or params.get("reference_image_url"))
    source_sku_id = clean_text(params.get("source_sku_id") or params.get("sku_id"))
    if not (record_id and image_url):
        return None
    return {
        "record_id": record_id,
        "source_sku_id": source_sku_id,
        "sku_id": source_sku_id,
        "sku": sku or source_sku_id,
        "sku_name": clean_text(params.get("sku_name") or sku or source_sku_id),
        "image_url": image_url,
        "sku_image_url": image_url,
        "price": params.get("price") or params.get("cost_price") or "",
        "cost_price": params.get("cost_price") or params.get("price") or "",
        "stock": params.get("stock", ""),
        "candidate": {"record_id": record_id, "business_key": clean_text(params.get("business_key") or sku)},
    }


def _export_fields(params: dict[str, Any], mode: str, embed_images: bool) -> list[str]:
    fields = params.get("fields") or params.get("columns")
    if isinstance(fields, str):
        fields = [item.strip() for item in re.split(r"[,，、\s]+", fields) if item.strip()]
    if not isinstance(fields, list) or not fields:
        fields = list(DEFAULT_EXPORT_FIELDS)
    clean_fields = [clean_text(item) for item in fields if clean_text(item)]
    if mode == "internal" and "customer_price" in clean_fields and not params.get("include_customer_price"):
        clean_fields = [field for field in clean_fields if field != "customer_price"]
    if embed_images and "sku_image" not in clean_fields:
        clean_fields.insert(0, "sku_image")
    return clean_fields


def _resolve_output_path(params: dict[str, Any], capture: dict[str, Any]) -> Path:
    raw_output_path = clean_text(params.get("output_path") or params.get("workbook_path"))
    if raw_output_path:
        return Path(raw_output_path)
    raw_dir = clean_text(params.get("output_dir") or params.get("export_dir") or "")
    output_dir = Path(raw_dir) if raw_dir else exports_dir()
    file_name = _safe_file_stem(params.get("filename") or capture.get("title") or "reference-product") + "-" + _timestamp_for_file() + ".xlsx"
    return output_dir / file_name


def _margin_percent(params: dict[str, Any]) -> Decimal:
    if params.get("margin_percent") not in (None, ""):
        return _decimal(params.get("margin_percent"))
    value = params.get("margin_rate")
    if value not in (None, ""):
        decimal_value = _decimal(value)
        if Decimal("-1") < decimal_value < Decimal("1"):
            return decimal_value * Decimal("100")
        return decimal_value
    return _decimal(params.get("margin") or params.get("markup_percent") or 0)


def _first_customer_price_text(rows: list[dict[str, Any]]) -> str:
    for row in rows:
        text = clean_text(row.get("customer_price_text"))
        if text:
            return text
    return "-"


def _export_rows(
    capture: dict[str, Any],
    match: dict[str, Any] | None,
    fields: list[str],
    mode: str,
    margin_percent: Decimal,
    decimal_places: int,
) -> list[dict[str, Any]]:
    match_by_sku = {}
    if match:
        for item in match.get("matches", []):
            match_by_sku[(clean_text(item.get("sku_id")), clean_text(item.get("sku_name")))] = item
    rows = []
    for index, sku in enumerate(capture.get("skus", []), start=1):
        cost_price = _decimal_or_blank(sku.get("cost_price") if sku.get("cost_price") not in (None, "") else sku.get("price"))
        customer_price = ""
        if cost_price != "":
            customer_price = _quantize(cost_price * (Decimal("1") + margin_percent / Decimal("100")), decimal_places)
        match_item = match_by_sku.get((clean_text(sku.get("sku_id")), clean_text(sku.get("sku_name")))) or {}
        candidate = match_item.get("best_candidate") if isinstance(match_item.get("best_candidate"), dict) else {}
        row = {
            "index": index,
            "sku_name": sku.get("sku_name") or "",
            "sku_id": sku.get("sku_id") or sku.get("source_sku_id") or "",
            "source_sku_id": sku.get("source_sku_id") or sku.get("sku_id") or "",
            "sku": sku.get("sku") or sku.get("sku_id") or "",
            "spec": sku.get("spec") or "",
            "cost_price": float(cost_price) if isinstance(cost_price, Decimal) else "",
            "cost_price_text": f"{_quantize(cost_price, decimal_places):.{decimal_places}f}" if isinstance(cost_price, Decimal) else "",
            "customer_price": float(customer_price) if customer_price != "" else "",
            "customer_price_text": f"{customer_price:.{decimal_places}f}" if customer_price != "" else "",
            "margin_rate": float(margin_percent / Decimal("100")),
            "price": sku.get("price") or "",
            "stock": sku.get("stock", ""),
            "sku_image": sku.get("image_url") or sku.get("sku_image_url") or "",
            "sku_image_url": sku.get("image_url") or sku.get("sku_image_url") or "",
            "image_url": sku.get("image_url") or sku.get("sku_image_url") or "",
            "source_url": capture.get("source_url") or "",
            "matched_record": candidate.get("title") or "",
        }
        rows.append(row)
    return rows


def _decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value).replace("%", "").strip() or "0")
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _decimal_or_blank(value: Any) -> Decimal | str:
    text = clean_text(value)
    if not text:
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return ""
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return ""


def _quantize(value: Decimal, decimal_places: int) -> Decimal:
    unit = Decimal("1") if decimal_places <= 0 else Decimal("1").scaleb(-decimal_places)
    return value.quantize(unit, rounding=ROUND_HALF_UP)


def _download_export_images(capture: dict[str, Any], rows: list[dict[str, Any]]) -> dict[int, str]:
    capture_image_dir = images_dir() / clean_text(capture.get("capture_id") or "latest")
    capture_image_dir.mkdir(parents=True, exist_ok=True)
    downloaded = {}
    for index, row in enumerate(rows, start=1):
        url = clean_text(row.get("sku_image_url"))
        if not url:
            continue
        try:
            file_path = _download_image(url, capture_image_dir, index)
        except Exception:
            continue
        if file_path:
            downloaded[index] = str(file_path)
    return downloaded


def _download_image(url: str, target_dir: Path, index: int) -> Path:
    if url.startswith("data:image/"):
        header, encoded = url.split(",", 1)
        mime = header.split(";", 1)[0].removeprefix("data:")
        ext = mimetypes.guess_extension(mime) or ".png"
        raw = base64.b64decode(encoded)
    elif url.startswith("file:///"):
        source = Path(url.replace("file:///", "").replace("/", "\\"))
        ext = source.suffix if source.suffix.lower() in SUPPORTED_IMAGE_EXTENSIONS else ".jpg"
        target = target_dir / f"sku-{index}{ext}"
        shutil.copyfile(source, target)
        return target
    else:
        raw, response = fetch_bytes(url, timeout=20)
        content_type = response.headers.get("Content-Type", "") if hasattr(response, "headers") else ""
        ext = mimetypes.guess_extension(content_type.split(";", 1)[0].strip()) if content_type else ""
        if not ext:
            ext = Path(urlparse(url).path).suffix
        if ext.lower() not in SUPPORTED_IMAGE_EXTENSIONS:
            ext = ".jpg"
    target = target_dir / f"sku-{index}{ext}"
    target.write_bytes(raw)
    return target


def _write_export_workbook(output_path: Path, fields: list[str], rows: list[dict[str, Any]], images: dict[int, str], embed_images: bool) -> None:
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as OpenpyxlImage
    except ImportError as exc:
        raise ReferenceProductError("缺少 openpyxl，无法导出 Excel。") from exc

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "参考商品"
    for column_index, field in enumerate(fields, start=1):
        sheet.cell(row=1, column=column_index, value=FIELD_LABELS.get(field, field))
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = 18

    image_column = fields.index("sku_image") + 1 if "sku_image" in fields else 0
    image_import_ok = True
    for row_index, row in enumerate(rows, start=2):
        for column_index, field in enumerate(fields, start=1):
            if field == "sku_image" and embed_images and row_index - 1 in images:
                sheet.cell(row=row_index, column=column_index, value="")
                continue
            sheet.cell(row=row_index, column=column_index, value=row.get(field, ""))
        if image_column and embed_images and row_index - 1 in images:
            try:
                image = OpenpyxlImage(images[row_index - 1])
                image.width = 90
                image.height = 90
                sheet.add_image(image, f"{openpyxl.utils.get_column_letter(image_column)}{row_index}")
                sheet.row_dimensions[row_index].height = 72
            except Exception:
                image_import_ok = False
                sheet.cell(row=row_index, column=image_column, value=images[row_index - 1])
    if embed_images and not image_import_ok:
        sheet.cell(row=len(rows) + 3, column=1, value="部分图片未能嵌入，已保留本地图片路径。")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def _safe_file_stem(value: Any) -> str:
    text = clean_text(value) or "reference-product"
    text = re.sub(r"[<>:\"/\\|?*\u0000-\u001F]+", "", text).strip(". ")
    return text[:80] or "reference-product"


def _image_markdown(url: Any) -> str:
    clean = clean_text(url)
    return f"![图]({clean})" if clean else "-"


def _cache_size(root: Path) -> tuple[int, int]:
    total = 0
    count = 0
    for item in root.rglob("*"):
        if not item.is_file():
            continue
        try:
            total += item.stat().st_size
            count += 1
        except OSError:
            continue
    return total, count


def _format_bytes(value: int) -> str:
    number = float(value)
    for unit in ("B", "KB", "MB", "GB"):
        if number < 1024 or unit == "GB":
            return f"{number:.1f}{unit}" if unit != "B" else f"{int(number)}B"
        number /= 1024
    return f"{number:.1f}GB"


def _remove_empty_dirs(root: Path) -> None:
    for item in sorted((path for path in root.rglob("*") if path.is_dir()), key=lambda path: len(path.parts), reverse=True):
        try:
            item.rmdir()
        except OSError:
            continue
