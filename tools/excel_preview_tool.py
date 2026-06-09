from __future__ import annotations

import re
import shutil
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, BinaryIO

from core.local_runtime import local_runtime_dir
from tools.excel_table_classifier import (
    FIELD_CATALOG_VERSION,
    classify_table,
    standard_field_catalog,
    summarize_classifications,
    table_category_catalog,
)


ALLOWED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MAX_PREVIEW_ROWS = 20
DEFAULT_HEADER_SCAN_ROWS = 20
HEADER_KEYWORDS = (
    "货号",
    "品名",
    "产品",
    "商品",
    "包装",
    "厂价",
    "单价",
    "成本",
    "报价",
    "价格",
    "规格",
    "数量",
    "装箱",
    "外箱",
    "毛重",
    "净重",
    "毛净重",
    "尺寸",
    "条码",
    "客户",
    "供应商",
    "工厂",
    "厂家",
    "订单",
    "库存",
    "仓库",
    "采购",
    "金额",
    "日期",
    "联系人",
    "联系电话",
)


def imports_dir() -> Path:
    return local_runtime_dir() / "imports"


def _clean_filename(filename: str) -> str:
    name = Path(filename or "uploaded.xlsx").name.strip()
    stem = Path(name).stem or "uploaded"
    suffix = Path(name).suffix.lower()
    stem = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", stem, flags=re.UNICODE).strip("._") or "uploaded"
    return f"{stem}{suffix}"


def _validate_excel_path(path: Path) -> None:
    if path.suffix.lower() not in ALLOWED_EXCEL_EXTENSIONS:
        raise ValueError("只支持 .xlsx、.xlsm、.xltx、.xltm 格式的 Excel 文件。")


def _excel_files_in_dir(path: Path) -> list[Path]:
    return sorted(
        item
        for item in path.iterdir()
        if item.is_file()
        and item.suffix.lower() in ALLOWED_EXCEL_EXTENSIONS
        and not item.name.startswith("~$")
    )


def save_uploaded_excel(filename: str, stream: BinaryIO) -> Path:
    clean_name = _clean_filename(filename)
    candidate = Path(clean_name)
    _validate_excel_path(candidate)

    target_dir = imports_dir()
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S_%f")
    target = target_dir / f"{timestamp}_{uuid.uuid4().hex[:8]}_{clean_name}"
    with target.open("wb") as output:
        shutil.copyfileobj(stream, output)
    return target


def _load_openpyxl():
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法读取 Excel 文件。") from exc
    return openpyxl


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def _row_values(ws, row: int) -> list[str]:
    return [_cell_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]


def _trim_row(row: list[str]) -> list[str]:
    clean = list(row)
    while clean and not clean[-1]:
        clean.pop()
    return clean


def _header_keyword_score(values: list[str]) -> int:
    score = 0
    for value in values:
        clean = re.sub(r"\s+", "", value)
        if not clean:
            continue
        score += sum(1 for keyword in HEADER_KEYWORDS if keyword in clean)
    return score


def _numeric_cell_count(values: list[str]) -> int:
    count = 0
    for value in values:
        clean = value.replace(",", "").replace("%", "").strip()
        if clean and re.fullmatch(r"[-+]?\d+(?:\.\d+)?", clean):
            count += 1
    return count


def _find_header_row(ws) -> int:
    best_row = 1
    best_score = -1
    max_scan = min(ws.max_row, DEFAULT_HEADER_SCAN_ROWS)
    for row_index in range(1, max_scan + 1):
        values = _row_values(ws, row_index)
        non_empty = sum(1 for value in values if value)
        unique_values = len({value for value in values if value})
        keyword_score = _header_keyword_score(values)
        numeric_count = _numeric_cell_count(values)
        score = keyword_score * 12 + non_empty * 2 + unique_values - numeric_count * 3
        if non_empty >= 2 and score > best_score:
            best_score = score
            best_row = row_index
    return best_row


def _preview_sheet(ws, preview_rows: int) -> dict[str, Any]:
    header_row = _find_header_row(ws)
    headers = _trim_row(_row_values(ws, header_row))
    if not headers:
        headers = [f"列{index}" for index in range(1, ws.max_column + 1)]
    normalized_headers = [
        header if header else f"未命名列{index}"
        for index, header in enumerate(headers, start=1)
    ]
    width = len(normalized_headers)
    leading_rows: list[list[str]] = []
    for row_index in range(1, header_row):
        row = _trim_row(_row_values(ws, row_index))
        if any(row):
            leading_rows.append(row)
    rows: list[list[str]] = []
    for row_index in range(header_row + 1, ws.max_row + 1):
        row = _row_values(ws, row_index)[:width]
        if not any(row):
            continue
        row += [""] * (width - len(row))
        rows.append(row)
        if len(rows) >= preview_rows:
            break
    classification = classify_table(normalized_headers, rows)
    return {
        "name": ws.title,
        "header_row": header_row,
        "headers": normalized_headers,
        "leading_rows": leading_rows,
        "rows": rows,
        "preview_row_count": len(rows),
        "total_rows": ws.max_row,
        "total_columns": ws.max_column,
        "classification": classification,
    }


def preview_excel_file(path: str | Path, preview_rows: int = MAX_PREVIEW_ROWS) -> dict[str, Any]:
    workbook_path = Path(path)
    if workbook_path.exists() and workbook_path.is_dir():
        files = _excel_files_in_dir(workbook_path)
        if not files:
            raise FileNotFoundError(f"文件夹中没有可读取的 Excel 文件：{workbook_path}")
        choices = "、".join(item.name for item in files[:10])
        raise ValueError(f"提供的是文件夹，请指定具体 Excel 文件。可选文件：{choices}")
    _validate_excel_path(workbook_path)
    if not workbook_path.exists() or not workbook_path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在：{workbook_path}")

    safe_preview_rows = max(1, min(int(preview_rows or MAX_PREVIEW_ROWS), MAX_PREVIEW_ROWS))
    openpyxl = _load_openpyxl()
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheets = [_preview_sheet(ws, safe_preview_rows) for ws in workbook.worksheets]
    finally:
        workbook.close()

    classification_summary = summarize_classifications(sheets)
    return {
        "ok": True,
        "kind": "excel_preview",
        "title": "Excel 通用表格识别预览",
        "summary": {
            "文件": workbook_path.name,
            "保存路径": str(workbook_path),
            "工作表数": len(sheets),
            "预览行数": safe_preview_rows,
            "已识别工作表": classification_summary["status_counts"].get("recognized", 0),
            "未识别工作表": classification_summary["status_counts"].get("unknown", 0),
            "待确认建议": classification_summary["change_proposal_count"],
        },
        "data": {
            "file_name": workbook_path.name,
            "saved_path": str(workbook_path),
            "sheet_count": len(sheets),
            "sheets": sheets,
            "classification_summary": classification_summary,
            "field_catalog_version": FIELD_CATALOG_VERSION,
            "standard_fields": standard_field_catalog(),
            "table_categories": table_category_catalog(),
        },
    }


def _resolve_excel_path(params: dict[str, Any]) -> Path:
    raw_path = params.get("path") or params.get("source_path") or params.get("workbook_path")
    filename = params.get("filename") or params.get("file_name") or params.get("name")
    if not raw_path:
        if filename:
            return Path(str(filename))
        raise ValueError("请提供 path、source_path 或 workbook_path。")

    path = Path(str(raw_path))
    if path.exists() and path.is_dir():
        if filename:
            return path / Path(str(filename)).name
        files = _excel_files_in_dir(path)
        if len(files) == 1:
            return files[0]
        if not files:
            raise FileNotFoundError(f"文件夹中没有可读取的 Excel 文件：{path}")
        choices = "、".join(item.name for item in files[:10])
        raise ValueError(f"source_path 是文件夹，请提供 filename 或 workbook_path。可选文件：{choices}")
    return path


def preview_excel(params: dict[str, Any]) -> dict[str, Any]:
    params = params or {}
    try:
        path = _resolve_excel_path(params)
        return preview_excel_file(str(path), int(params.get("preview_rows") or MAX_PREVIEW_ROWS))
    except Exception as exc:
        return {
            "ok": False,
            "kind": "excel_preview_error",
            "title": "Excel 读取失败",
            "message": str(exc),
        }
