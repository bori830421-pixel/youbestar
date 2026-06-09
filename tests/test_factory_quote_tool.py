import unittest
from pathlib import Path
from typing import Any
from unittest.mock import patch

import openpyxl

from tools.factory_quote_tool import import_factory_quotes, library_status, query_factory_quote, run


TEST_DATA_DIR = Path(__file__).resolve().parents[1] / "data"


HEADERS = [
    "中文包装",
    "产品图片",
    "货号",
    "品名",
    "包装",
    "装箱数量",
    "品牌价",
    "单价    (元/件)",
    "外箱规格(cm)",
    "包装规格(cm)",
    "产品尺寸（cm）",
    "毛重\n(公斤)",
    "净重\n(公斤)",
    "条码",
    "备注",
]


def build_workbook(factory: str, price: float = 10.0, brand: str = "", title: str | None = None):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws.merge_cells("A1:N2")
    title_text = title if title is not None else f"{factory}\n产品报价表\n业务员：潘小姐 13502760103"
    ws["A1"] = title_text
    headers = list(HEADERS)
    if brand:
        headers.insert(4, "品牌")
    for index, header in enumerate(headers, start=1):
        ws.cell(3, index).value = header
    ws["A4"] = "桌游"
    ws["C4"] = "QQL701A"
    ws["D4"] = "大盒五子棋"
    offset = 1 if brand else 0
    if brand:
        ws["E4"] = brand
    ws.cell(4, 5 + offset).value = "彩盒"
    ws.cell(4, 6 + offset).value = 36
    ws.cell(4, 7 + offset).value = price
    ws.cell(4, 8 + offset).value = price * 36
    ws.cell(4, 9 + offset).value = "63*33*38"
    ws.cell(4, 10 + offset).value = "28.9*14.5*3.8"
    ws.cell(4, 11 + offset).value = "28.9*28.9*2.3"
    ws.cell(4, 12 + offset).value = 23
    ws.cell(4, 13 + offset).value = 21
    ws.cell(4, 14 + offset).value = "6956325507011"
    return workbook


def build_complex_packaging_workbook(
    weight_value: Any = "23/21",
    single_gross: Any = "638.9g",
    single_net: Any = "583.3g",
    single_weight: Any = None,
    product_spec: str = "27*39.5*5.5",
):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws["A1"] = "吉贵（原中意）玩具厂报价表\n联系人：朱召深 13923663353"
    headers = [
        "中文包装",
        "货号",
        "品名",
        "包装",
        "装箱数量",
        "内盒",
        "内盒规格",
        "品牌价",
        "外箱规格(CM)",
        "棋盘规格",
        "毛/净重(KG)",
        "单品毛重",
        "单品净重",
        "单品克重",
        "快递包装重量",
        "条码",
    ]
    for index, header in enumerate(headers, start=1):
        ws.cell(3, index).value = header
    values = [
        "棋类",
        "8001",
        "英.阿文.序列棋牌/纸棋盘",
        "天地盒",
        36,
        4,
        "30*20*15",
        23.3,
        "63*33*38",
        product_spec,
        weight_value,
        single_gross,
        single_net,
        single_weight,
        "650g",
        "6956325507011",
    ]
    for index, value in enumerate(values, start=1):
        ws.cell(4, index).value = value
    return workbook


def build_zhongyi_workbook():
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Sheet1"
    ws["A1"] = "吉贵（原中意）玩具厂报价表"
    ws["A3"] = "联系人：朱召深 联系电话：13923663353 83252311 传真：85631232"
    headers = [
        "",
        "货号",
        "包装/品名",
        "包装",
        "厂价(RMB)",
        "",
        "包装规格",
        "单品克重",
        "棋盘规格",
        "数量(只)",
        "内盒",
        "外箱规格(CM)",
        "毛净重(KG)",
    ]
    for index, header in enumerate(headers, start=1):
        ws.cell(4, index).value = header
    values = [
        "",
        "8001",
        "英.阿文.序列棋牌/纸棋盘",
        "天地盒",
        23.3,
        "",
        "27*39.5*5.5",
        650,
        "27*39.5*5.5",
        36,
        4,
        "63*33*38",
        "23/21",
    ]
    for index, value in enumerate(values, start=1):
        ws.cell(5, index).value = value
    return workbook


class FakeOpenpyxl:
    def __init__(self, workbooks: dict[str, Any]):
        self.workbooks = workbooks

    def load_workbook(self, path, data_only=False):
        return self.workbooks[str(path)]


class FactoryQuoteToolTest(unittest.TestCase):
    def make_library_path(self) -> str:
        TEST_DATA_DIR.mkdir(exist_ok=True)
        name = self.id().rsplit(".", 1)[-1]
        return str(TEST_DATA_DIR / f"test_factory_quote_{name}.sqlite3")

    def with_workbooks(self, workbooks: dict[str, Any], func, params):
        paths = [Path(path) for path in workbooks]
        with (
            patch("tools.factory_quote_tool._resolve_source_paths", return_value=paths),
            patch("tools.factory_quote_tool._load_openpyxl", return_value=FakeOpenpyxl(workbooks)),
        ):
            return func(params)

    def test_imports_excel_quote_rows(self):
        path = r"D:\工厂报价\quote.xlsx"
        result = self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司")},
            import_factory_quotes,
            {"workbook_path": path, "library_path": self.make_library_path()},
        )

        self.assertIs(result["ok"], True)
        self.assertEqual(result["kind"], "factory_quote_import")
        self.assertEqual(result["summary"]["产品数"], 1)
        self.assertIn("已写入本地报价资料库", result["summary"]["状态"])
        self.assertEqual(result["data"]["products"][0]["sku"], "QQL701A")
        self.assertEqual(result["data"]["products"][0]["factory_contact"]["business_contact"], "潘小姐")
        self.assertEqual(result["data"]["products"][0]["factory_contact"]["business_phone"], "13502760103")

    def test_import_blocks_when_factory_and_brand_are_unidentified(self):
        path = r"D:\工厂报价\unknown.xlsx"
        library_path = self.make_library_path()
        result = self.with_workbooks(
            {path: build_workbook("", title="产品报价表\n联系人：潘小姐 13502760103")},
            import_factory_quotes,
            {"workbook_path": path, "library_path": library_path},
        )
        status = library_status({"library_path": library_path})

        self.assertIs(result["ok"], False)
        self.assertEqual(result["kind"], "factory_quote_import_identity_required")
        self.assertTrue(result["data"]["requires_confirmation"])
        self.assertEqual(result["summary"]["状态"], "未写入数据库")
        self.assertEqual(status["summary"]["产品数"], 0)

    def test_import_accepts_confirmed_brand_when_factory_missing(self):
        path = r"D:\工厂报价\brand_only.xlsx"
        library_path = self.make_library_path()
        result = self.with_workbooks(
            {path: build_workbook("", title="产品报价表\n联系人：潘小姐 13502760103")},
            import_factory_quotes,
            {"workbook_path": path, "library_path": library_path, "brand": "星河牌"},
        )
        query_result = query_factory_quote(
            {
                "library_path": library_path,
                "brand": "星河牌",
                "sku": "QQL701A",
            }
        )

        self.assertIs(result["ok"], True)
        self.assertEqual(result["summary"]["工厂数"], 0)
        self.assertEqual(result["summary"]["品牌数"], 1)
        self.assertEqual(result["data"]["products"][0]["brand"], "星河牌")
        self.assertIs(query_result["ok"], True)
        self.assertEqual(query_result["summary"]["品牌"], "星河牌")

    def test_import_stores_and_filters_brand_column(self):
        library_path = self.make_library_path()
        workbooks = {
            r"D:\工厂报价\brand_a.xlsx": build_workbook("潘多多（汕头）科教实业有限公司", price=10, brand="星河牌"),
            r"D:\工厂报价\brand_b.xlsx": build_workbook("潘多多（汕头）科教实业有限公司", price=12, brand="启明星"),
        }
        self.with_workbooks(workbooks, import_factory_quotes, {"source_path": r"D:\工厂报价", "library_path": library_path})

        result = query_factory_quote(
            {
                "library_path": library_path,
                "factory_name": "潘多多",
                "brand": "启明星",
                "sku": "QQL701A",
            }
        )

        self.assertIs(result["ok"], True)
        self.assertEqual(result["summary"]["品牌"], "启明星")
        self.assertEqual(result["summary"]["成本单价"], "12.00元")

    def test_query_by_factory_and_sku_returns_dimensions_and_costs(self):
        path = r"D:\工厂报价\quote.xlsx"
        result = self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=17.745)},
            query_factory_quote,
            {
                "workbook_path": path,
                "library_path": self.make_library_path(),
                "factory_name": "潘多多",
                "sku": "QQL701A",
            },
        )

        self.assertIs(result["ok"], True)
        self.assertNotIn("工厂", result["columns"])
        self.assertEqual(result["columns"][0], "SKU图")
        self.assertIn("箱规尺寸(cm)", result["columns"])
        self.assertIn("箱净重(kg)", result["columns"])
        self.assertIn("单品净重(g)", result["columns"])
        self.assertEqual(result["summary"]["货号"], "QQL701A")
        self.assertEqual(result["summary"]["产品尺寸"], "28.9*28.9*2.3cm")
        self.assertEqual(result["summary"]["箱规尺寸"], "63*33*38cm")
        self.assertEqual(result["summary"]["箱净重"], "21kg")
        self.assertEqual(result["summary"]["单品净重"], "约583.3g（箱重/装箱数换算）")
        self.assertEqual(result["summary"]["单品毛重"], "约638.9g（箱重/装箱数换算）")
        self.assertEqual(result["summary"]["快递包装重量"], "未记录")
        self.assertEqual(result["summary"]["成本单价"], "17.75元")
        self.assertEqual(result["summary"]["业务联系人"], "潘小姐")
        self.assertEqual(result["summary"]["业务联系电话"], "13502760103")
        self.assertEqual(result["data"]["contact"]["business_contact"], "潘小姐")
        self.assertEqual(result["data"]["contact"]["business_phone"], "13502760103")
        self.assertEqual(result["data"]["needs_disambiguation"], False)

    def test_contact_operation_returns_factory_business_contact(self):
        path = r"D:\工厂报价\quote.xlsx"
        library_path = self.make_library_path()
        self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=17.745)},
            import_factory_quotes,
            {"workbook_path": path, "library_path": library_path},
        )

        result = run(
            {
                "operation": "contact",
                "library_path": library_path,
                "factory_name": "潘多多",
            }
        )

        self.assertIs(result["ok"], True)
        self.assertEqual(result["kind"], "factory_contact")
        self.assertEqual(result["summary"]["业务联系人"], "潘小姐")
        self.assertEqual(result["summary"]["业务联系电话"], "13502760103")
        self.assertEqual(result["rows"][0][1], "潘小姐")
        self.assertEqual(result["rows"][0][2], "13502760103")

    def test_duplicate_sku_requires_factory_disambiguation(self):
        library_path = self.make_library_path()
        workbooks = {
            r"D:\工厂报价\factory_a.xlsx": build_workbook("潘多多（汕头）科教实业有限公司", price=10),
            r"D:\工厂报价\factory_b.xlsx": build_workbook("另一家工厂", price=12),
        }
        ambiguous = self.with_workbooks(workbooks, query_factory_quote, {"source_path": r"D:\工厂报价", "library_path": library_path, "sku": "QQL701A"})
        exact = self.with_workbooks(
            workbooks,
            query_factory_quote,
            {"source_path": r"D:\工厂报价", "library_path": library_path, "factory_name": "另一家", "sku": "QQL701A"},
        )

        self.assertIs(ambiguous["ok"], True)
        self.assertEqual(ambiguous["summary"]["匹配数量"], 2)
        self.assertTrue(ambiguous["data"]["needs_disambiguation"])
        self.assertEqual(exact["summary"]["工厂"], "另一家工厂")
        self.assertEqual(exact["summary"]["成本单价"], "12.00元")

    def test_quantity_pricing_uses_margin_tax_and_freight(self):
        path = r"D:\工厂报价\quote.xlsx"
        result = self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=10)},
            query_factory_quote,
            {
                "workbook_path": path,
                "library_path": self.make_library_path(),
                "factory_name": "潘多多",
                "sku": "QQL701A",
                "quantity": 100,
                "margin_rate": 0.2,
                "include_tax": True,
                "tax_rate": 0.13,
                "include_freight": True,
                "freight_fee": 50,
            },
        )

        calculation = result["data"]["calculation"]
        self.assertEqual(calculation["estimated_cartons"], 3)
        self.assertEqual(calculation["quote_unit_price"], 13.56)
        self.assertEqual(calculation["quote_total"], 1406.0)
        self.assertEqual(result["summary"]["报价单价"], "13.56元")
        self.assertEqual(result["summary"]["报价总额"], "1406.00元")

    def test_price_display_uses_half_up_two_decimal_currency(self):
        path = r"D:\工厂报价\quote.xlsx"
        result = self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=15.225)},
            query_factory_quote,
            {
                "workbook_path": path,
                "library_path": self.make_library_path(),
                "factory_name": "潘多多",
                "sku": "QQL701A",
            },
        )

        self.assertEqual(result["summary"]["成本单价"], "15.23元")
        self.assertIn("成本单价(元)", result["columns"])
        self.assertEqual(result["rows"][0][result["columns"].index("成本单价(元)")], "15.23")

    def test_import_once_then_query_reads_from_library_without_excel(self):
        path = r"D:\工厂报价\quote.xlsx"
        library_path = self.make_library_path()
        self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=17.745)},
            import_factory_quotes,
            {"workbook_path": path, "library_path": library_path},
        )

        with patch("tools.factory_quote_tool.load_quote_products", side_effect=AssertionError("should read sqlite library")):
            result = query_factory_quote(
                {
                    "library_path": library_path,
                    "factory_name": "潘多多",
                    "sku": "QQL701A",
                }
            )

        status = library_status({"library_path": library_path})

        self.assertIs(result["ok"], True)
        self.assertEqual(result["summary"]["成本单价"], "17.75元")
        self.assertEqual(result["data"]["source"], "library")
        self.assertEqual(status["summary"]["产品数"], 1)

    def test_creates_pending_image_asset_candidate_for_1688_link(self):
        result = run(
            {
                "operation": "1688",
                "library_path": self.make_library_path(),
                "factory_name": "潘多多",
                "sku": "QQL701A",
                "source_url": "https://detail.1688.com/offer/123.html",
            }
        )

        self.assertIs(result["ok"], True)
        self.assertEqual(result["kind"], "image_asset_candidate")
        self.assertEqual(result["data"]["source_type"], "1688_product_page")
        self.assertEqual(result["data"]["status"], "pending_confirmation")

    def test_bind_image_then_quote_query_returns_thumbnail(self):
        path = r"D:\工厂报价\quote.xlsx"
        library_path = self.make_library_path()
        self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=17.745)},
            import_factory_quotes,
            {"workbook_path": path, "library_path": library_path},
        )

        image_url = "https://cbu01.alicdn.com/img/ibank/example.jpg"
        bind_result = run(
            {
                "operation": "bind_image",
                "library_path": library_path,
                "factory_name": "潘多多",
                "sku": "QQL701A",
                "image_url": image_url,
            }
        )
        quote_result = query_factory_quote(
            {
                "library_path": library_path,
                "factory_name": "潘多多",
                "sku": "QQL701A",
            }
        )

        self.assertIs(bind_result["ok"], True)
        self.assertEqual(bind_result["kind"], "image_asset_binding")
        self.assertEqual(bind_result["data"]["image_asset"]["status"], "confirmed")
        self.assertEqual(bind_result["data"]["image_asset"]["image_type"], "sku_image")
        self.assertEqual(bind_result["summary"]["图片类型"], "SKU图")
        self.assertEqual(quote_result["columns"][0], "SKU图")
        self.assertNotIn("工厂", quote_result["columns"])
        self.assertEqual(quote_result["rows"][0][0], f"![SKU图]({image_url})")
        self.assertEqual(quote_result["summary"]["SKU图"], "已绑定")
        self.assertEqual(quote_result["summary"]["实拍图数量"], 0)
        self.assertEqual(quote_result["data"]["sku_images"][0]["original_image_url"], image_url)

    def test_real_photos_are_multiple_and_do_not_replace_sku_image(self):
        path = r"D:\工厂报价\quote.xlsx"
        library_path = self.make_library_path()
        self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=17.745)},
            import_factory_quotes,
            {"workbook_path": path, "library_path": library_path},
        )

        sku_image = "https://cbu01.alicdn.com/img/ibank/sku.jpg"
        real_photo_a = "https://cbu01.alicdn.com/img/ibank/real-a.jpg"
        real_photo_b = "https://cbu01.alicdn.com/img/ibank/real-b.jpg"
        for image_url, image_type in (
            (sku_image, "sku_image"),
            (real_photo_a, "real_photo"),
            (real_photo_b, "实拍图"),
        ):
            result = run(
                {
                    "operation": "bind_image",
                    "library_path": library_path,
                    "factory_name": "潘多多",
                    "sku": "QQL701A",
                    "image_url": image_url,
                    "image_type": image_type,
                }
            )
            self.assertIs(result["ok"], True)

        quote_result = query_factory_quote(
            {
                "library_path": library_path,
                "factory_name": "潘多多",
                "sku": "QQL701A",
            }
        )

        self.assertEqual(quote_result["rows"][0][0], f"![SKU图]({sku_image})")
        self.assertEqual(quote_result["summary"]["实拍图数量"], 2)
        self.assertEqual(len(quote_result["data"]["real_photos"][0]), 2)
        self.assertEqual({photo["original_image_url"] for photo in quote_result["data"]["real_photos"][0]}, {real_photo_a, real_photo_b})

    def test_rebinding_sku_image_replaces_previous_sku_image_only(self):
        path = r"D:\工厂报价\quote.xlsx"
        library_path = self.make_library_path()
        self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=17.745)},
            import_factory_quotes,
            {"workbook_path": path, "library_path": library_path},
        )

        first_sku_image = "https://cbu01.alicdn.com/img/ibank/sku-old.jpg"
        second_sku_image = "https://cbu01.alicdn.com/img/ibank/sku-new.jpg"
        real_photo = "https://cbu01.alicdn.com/img/ibank/real.jpg"
        for params in (
            {"image_url": first_sku_image},
            {"image_url": real_photo, "image_type": "real_photo"},
            {"image_url": second_sku_image},
        ):
            result = run(
                {
                    "operation": "bind_image",
                    "library_path": library_path,
                    "factory_name": "潘多多",
                    "sku": "QQL701A",
                    **params,
                }
            )
            self.assertIs(result["ok"], True)

        quote_result = query_factory_quote(
            {
                "library_path": library_path,
                "factory_name": "潘多多",
                "sku": "QQL701A",
            }
        )

        self.assertEqual(quote_result["rows"][0][0], f"![SKU图]({second_sku_image})")
        self.assertEqual(len(quote_result["data"]["sku_images"]), 1)
        self.assertEqual(quote_result["data"]["sku_images"][0]["original_image_url"], second_sku_image)
        self.assertEqual(len(quote_result["data"]["real_photos"][0]), 1)
        self.assertEqual(quote_result["data"]["real_photos"][0][0]["original_image_url"], real_photo)

    def test_single_product_weights_are_derived_from_carton_weights_in_grams(self):
        path = r"D:\工厂报价\quote.xlsx"
        result = self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=17.745)},
            query_factory_quote,
            {
                "workbook_path": path,
                "library_path": self.make_library_path(),
                "factory_name": "潘多多",
                "sku": "QQL701A",
            },
        )

        net_index = result["columns"].index("单品净重(g)")
        gross_index = result["columns"].index("单品毛重(g)")
        self.assertEqual(result["rows"][0][net_index], "约583.3g（箱重/装箱数换算）")
        self.assertEqual(result["rows"][0][gross_index], "约638.9g（箱重/装箱数换算）")
        self.assertEqual(result["data"]["products"][0]["net_weight_kg"], 21)
        self.assertEqual(result["summary"]["箱净重"], "21kg")

    def test_import_parses_inner_box_product_spec_and_combined_weights(self):
        path = r"D:\工厂报价\complex.xlsx"
        result = self.with_workbooks(
            {path: build_complex_packaging_workbook()},
            query_factory_quote,
            {
                "workbook_path": path,
                "library_path": self.make_library_path(),
                "factory_name": "吉贵",
                "sku": "8001",
            },
        )

        self.assertIs(result["ok"], True)
        product = result["data"]["products"][0]
        self.assertEqual(product["factory_name"], "吉贵（原中意）玩具厂")
        self.assertEqual(product["inner_box_quantity"], 4)
        self.assertEqual(product["inner_box_size_cm"], "30*20*15")
        self.assertEqual(product["product_size_cm"], "27*39.5*5.5")
        self.assertEqual(product["carton_size_cm"], "63*33*38")
        self.assertEqual(product["gross_weight_kg"], 23)
        self.assertEqual(product["net_weight_kg"], 21)
        self.assertEqual(product["single_gross_weight_g"], 638.9)
        self.assertEqual(product["single_net_weight_g"], 583.3)
        self.assertEqual(product["shipping_packaged_weight_g"], 650)
        self.assertEqual(product["weight_text"], "23/21")
        self.assertEqual(result["summary"]["内盒数量"], 4)
        self.assertEqual(result["summary"]["内盒尺寸"], "30*20*15cm")
        self.assertEqual(result["summary"]["箱毛重"], "23kg")
        self.assertEqual(result["summary"]["箱净重"], "21kg")
        self.assertEqual(result["summary"]["单品毛重"], "638.9g")
        self.assertEqual(result["summary"]["单品净重"], "583.3g")
        self.assertEqual(result["summary"]["快递包装重量"], "650g")
        self.assertIn("重量原文", result["columns"])
        self.assertEqual(result["rows"][0][result["columns"].index("重量原文")], "23/21")
        self.assertEqual(result["rows"][0][result["columns"].index("含税")], "未知")

    def test_import_parses_zhongyi_real_header_aliases(self):
        path = r"D:\工厂资料\中意（吉贵）棋报价总1.xlsx"
        result = self.with_workbooks(
            {path: build_zhongyi_workbook()},
            query_factory_quote,
            {
                "workbook_path": path,
                "library_path": self.make_library_path(),
                "factory_name": "吉贵",
                "sku": "8001",
            },
        )

        self.assertIs(result["ok"], True)
        product = result["data"]["products"][0]
        self.assertEqual(product["factory_name"], "吉贵（原中意）玩具厂")
        self.assertEqual(product["sku"], "8001")
        self.assertEqual(product["product_name"], "英.阿文.序列棋牌/纸棋盘")
        self.assertEqual(product["cost_unit_price"], 23.3)
        self.assertEqual(product["pcs_per_carton"], 36)
        self.assertEqual(product["package_size_cm"], "27*39.5*5.5")
        self.assertEqual(product["product_size_cm"], "27*39.5*5.5")
        self.assertEqual(product["inner_box_quantity"], 4)
        self.assertEqual(product["carton_size_cm"], "63*33*38")
        self.assertEqual(product["gross_weight_kg"], 23)
        self.assertEqual(product["net_weight_kg"], 21)
        self.assertEqual(product["single_gross_weight_g"], 650)
        self.assertEqual(product["single_net_weight_g"], 650)
        self.assertEqual(result["summary"]["业务联系人"], "朱召深")
        self.assertEqual(result["summary"]["业务联系电话"], "13923663353")

    def test_import_parses_labeled_combined_carton_weights(self):
        path = r"D:\工厂报价\labeled_weight.xlsx"
        result = self.with_workbooks(
            {path: build_complex_packaging_workbook(weight_value="毛重23KG 净重21KG")},
            query_factory_quote,
            {
                "workbook_path": path,
                "library_path": self.make_library_path(),
                "factory_name": "吉贵",
                "sku": "8001",
            },
        )

        product = result["data"]["products"][0]
        self.assertEqual(product["gross_weight_kg"], 23)
        self.assertEqual(product["net_weight_kg"], 21)
        self.assertEqual(product["weight_text"], "毛重23KG 净重21KG")
        self.assertEqual(result["summary"]["重量原文"], "毛重23KG 净重21KG")

    def test_import_keeps_ambiguous_single_weight_text_without_carton_split(self):
        path = r"D:\工厂报价\ambiguous_weight.xlsx"
        result = self.with_workbooks(
            {
                path: build_complex_packaging_workbook(
                    weight_value="23KG",
                    single_gross=None,
                    single_net=None,
                    single_weight="420g",
                    product_spec="磁性折叠棋盘套装",
                )
            },
            query_factory_quote,
            {
                "workbook_path": path,
                "library_path": self.make_library_path(),
                "factory_name": "吉贵",
                "sku": "8001",
            },
        )

        product = result["data"]["products"][0]
        self.assertIsNone(product["gross_weight_kg"])
        self.assertIsNone(product["net_weight_kg"])
        self.assertEqual(product["weight_text"], "23KG")
        self.assertEqual(product["product_spec"], "磁性折叠棋盘套装")
        self.assertEqual(product["dimension_text"], "磁性折叠棋盘套装")
        self.assertEqual(product["product_size_cm"], "")
        self.assertEqual(result["summary"]["产品规格"], "磁性折叠棋盘套装")
        self.assertEqual(result["summary"]["尺寸原文"], "磁性折叠棋盘套装")
        self.assertEqual(result["summary"]["单品毛重"], "420g")
        self.assertEqual(result["summary"]["单品净重"], "420g")

    def test_manual_single_weights_are_saved_to_product_library_and_override_derived_values(self):
        path = r"D:\工厂报价\quote.xlsx"
        library_path = self.make_library_path()
        self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=17.745)},
            import_factory_quotes,
            {"workbook_path": path, "library_path": library_path},
        )

        update_result = run(
            {
                "operation": "update_weight",
                "library_path": library_path,
                "factory_name": "潘多多",
                "sku": "QQL701A",
                "single_net_weight_g": 350,
                "single_gross_weight_g": "370克",
                "shipping_packaged_weight_g": "0.39kg",
            }
        )
        quote_result = query_factory_quote(
            {
                "library_path": library_path,
                "factory_name": "潘多多",
                "sku": "QQL701A",
            }
        )

        self.assertIs(update_result["ok"], True)
        self.assertEqual(update_result["kind"], "product_weight_update")
        self.assertEqual(update_result["summary"]["单品净重"], "350g")
        self.assertEqual(update_result["summary"]["单品毛重"], "370g")
        self.assertEqual(update_result["summary"]["快递包装重量"], "390g")
        self.assertEqual(quote_result["summary"]["单品净重"], "350g")
        self.assertEqual(quote_result["summary"]["单品毛重"], "370g")
        self.assertEqual(quote_result["summary"]["快递包装重量"], "390g")
        self.assertEqual(quote_result["data"]["products"][0]["manual_specs"]["shipping_packaged_weight_g"], 390)

    def test_manual_specs_update_saves_corrected_dimensions_to_library(self):
        path = r"D:\工厂报价\quote.xlsx"
        library_path = self.make_library_path()
        self.with_workbooks(
            {path: build_workbook("潘多多（汕头）科教实业有限公司", price=17.745)},
            import_factory_quotes,
            {"workbook_path": path, "library_path": library_path},
        )

        update_result = run(
            {
                "operation": "update_specs",
                "library_path": library_path,
                "factory_name": "潘多多",
                "sku": "QQL701A",
                "product_size_cm": "28.5*15*3cm",
                "package_size_cm": "28.5*30*3",
            }
        )
        quote_result = query_factory_quote(
            {
                "library_path": library_path,
                "factory_name": "潘多多",
                "sku": "QQL701A",
            }
        )

        self.assertIs(update_result["ok"], True)
        self.assertEqual(update_result["kind"], "product_specs_update")
        self.assertEqual(update_result["summary"]["产品尺寸"], "28.5*15*3cm")
        self.assertEqual(update_result["summary"]["包装尺寸"], "28.5*30*3cm")
        self.assertEqual(quote_result["summary"]["产品尺寸"], "28.5*15*3cm")
        self.assertEqual(quote_result["summary"]["包装尺寸"], "28.5*30*3cm")
        self.assertEqual(quote_result["data"]["products"][0]["product_size_cm"], "28.5*15*3")


if __name__ == "__main__":
    unittest.main()
