import unittest
import os
import tempfile
from pathlib import Path
from unittest.mock import patch

from fastapi import HTTPException
from fastapi.testclient import TestClient
import openpyxl

import server
from core.agent_state import AgentResult
from memory.memory import Memory


class FakeLLM:
    pass


class ServerChatRuntimeTest(unittest.TestCase):
    def setUp(self):
        self.original_memory = server.memory
        server.memory = Memory()

    def tearDown(self):
        server.memory = self.original_memory

    def test_chat_uses_owned_agent_runtime_by_default(self):
        self.assertTrue(server.USE_AGENT_RUNTIME)

    def test_chat_request_accepts_thread_id_for_owned_runtime(self):
        request = server.ChatRequest(message="你好", threadId="chat-123")

        self.assertEqual(request.threadId, "chat-123")

    def test_chat_request_accepts_tool_and_skill_switches(self):
        request = server.ChatRequest(
            message="你好",
            allowChat=True,
            allowTools=False,
            allowSkills=True,
            allowSelfEvolution=True,
        )

        self.assertTrue(request.allowChat)
        self.assertFalse(request.allowTools)
        self.assertTrue(request.allowSkills)
        self.assertTrue(request.allowSelfEvolution)

    def test_excel_preview_upload_endpoint_reads_all_sheets(self):
        workbook = openpyxl.Workbook()
        first = workbook.active
        first.title = "报价表"
        first.append(["货号", "品名"])
        first.append(["QQL701A", "大盒五子棋"])
        second = workbook.create_sheet("联系人")
        second.append(["工厂", "业务员"])
        second.append(["潘多多", "潘小姐"])

        from io import BytesIO

        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        client = TestClient(server.app)
        response = client.post(
            "/files/excel/preview?filename=quote.xlsx",
            content=buffer.getvalue(),
            headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual([sheet["name"] for sheet in data["data"]["sheets"]], ["报价表", "联系人"])
        self.assertEqual(data["data"]["sheets"][0]["headers"], ["货号", "品名"])
        self.assertEqual(data["data"]["sheets"][1]["rows"][0], ["潘多多", "潘小姐"])

    def test_excel_feedback_endpoint_persists_user_correction(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            local_home = Path(temp_dir) / "YoubestarLocal"
            with patch.dict(os.environ, {"YOUBESTAR_LOCAL_HOME": str(local_home)}):
                client = TestClient(server.app)
                feedback_response = client.post(
                    "/files/excel/feedback",
                    json={
                        "headers": ["名称", "数量", "备注"],
                        "sheet_name": "修正表",
                        "category": "inventory",
                        "field_mappings": {"名称": "product_name"},
                        "scope": "template",
                    },
                )
                self.assertEqual(feedback_response.status_code, 200)
                self.assertTrue(feedback_response.json()["ok"])

                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "修正表"
                sheet.append(["名称", "数量", "备注"])
                sheet.append(["积木A", 3, "样例"])

                from io import BytesIO

                buffer = BytesIO()
                workbook.save(buffer)
                workbook.close()

                preview_response = client.post(
                    "/files/excel/preview?filename=feedback.xlsx",
                    content=buffer.getvalue(),
                    headers={"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                )
                self.assertEqual(preview_response.status_code, 200)
                classification = preview_response.json()["data"]["sheets"][0]["classification"]
                self.assertEqual(classification["category"], "inventory")
                self.assertTrue(classification["feedback"]["applied"])
                self.assertTrue((local_home / "data" / "excel_feedback.sqlite3").exists())

    def test_business_records_endpoints_upsert_and_query_local_database(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            local_home = Path(temp_dir) / "YoubestarLocal"
            with patch.dict(os.environ, {"YOUBESTAR_LOCAL_HOME": str(local_home)}):
                client = TestClient(server.app)

                types_response = client.get("/business-records/types")
                self.assertEqual(types_response.status_code, 200)
                self.assertIn("customer", {item["record_type"] for item in types_response.json()["types"]})

                save_response = client.post(
                    "/business-records/upsert",
                    json={
                        "type": "customer",
                        "title": "汕头星河贸易",
                        "content": "联系人陈小姐，电话 13500000000",
                        "source": "unit-test",
                        "fields": {"customer_id": "C-001", "name": "汕头星河贸易", "contact": "陈小姐"},
                    },
                )
                self.assertEqual(save_response.status_code, 200)
                saved = save_response.json()
                self.assertTrue(saved["ok"])
                self.assertEqual(saved["record"]["business_key"], "C-001")
                self.assertTrue(saved["record"]["source_ip"])

                query_response = client.post("/business-records/query", json={"type": "customer", "query": "星河"})
                self.assertEqual(query_response.status_code, 200)
                queried = query_response.json()
                self.assertEqual(queried["summary"]["匹配数量"], 1)
                self.assertEqual(queried["records"][0]["fields"]["contact"], "陈小姐")
                self.assertTrue((local_home / "data" / "business_records.sqlite3").exists())

    def test_run_agent_runtime_returns_chat_response_shape(self):
        result = AgentResult(
            reply="你好，我在。",
            model_reply="Thought: 用户问候。\nAction: none\nParams: {}\nResponse: 你好，我在。",
            thought="用户问候。",
            action="none",
            params={},
            action_result="无操作",
            action_payload=None,
            response="你好，我在。",
            runtime_nodes=["direct_chat", "finalize"],
            thread_id="default",
            step_count=2,
        )

        with patch.object(server.agent_runtime, "run", return_value=result) as run_mock:
            response = server.run_agent_runtime(
                FakeLLM(),
                "你好",
                True,
                allow_tools=False,
                allow_skills=True,
                allow_self_evolution=True,
                thread_id="chat-1",
            )

        self.assertEqual(response.reply, "你好，我在。")
        self.assertEqual(response.action, "none")
        self.assertEqual(response.action_result, "无操作")
        self.assertIsNone(response.action_payload)
        self.assertEqual(response.interactions, [])
        self.assertEqual(response.response, "你好，我在。")
        self.assertIsNone(response.memory_candidate)
        run_mock.assert_called_once_with(
            unittest.mock.ANY,
            server.memory,
            "你好",
            allow_chat=True,
            allow_tools=False,
            allow_skills=True,
            allow_self_evolution=True,
            thread_id="chat-1",
            history=None,
        )

    def test_run_agent_runtime_passes_structured_interactions_to_chat_response(self):
        interaction = {
            "kind": "reference_product_match_review",
            "title": "参考商品候选确认",
            "items": [],
        }
        payload = {
            "ok": True,
            "kind": "reference_product_match",
            "match_id": "match-1",
        }
        result = AgentResult(
            reply="候选已生成。",
            model_reply="",
            thought="",
            action="official.reference_product",
            params={"operation": "match"},
            action_result=str(payload),
            action_payload=payload,
            response="候选已生成。",
            interactions=[interaction],
            runtime_nodes=["execute", "finalize"],
            thread_id="chat-1",
            step_count=2,
        )

        with patch.object(server.agent_runtime, "run", return_value=result):
            response = server.run_agent_runtime(FakeLLM(), "生成候选", True, thread_id="chat-1")

        self.assertEqual(response.action_payload, payload)
        self.assertEqual(response.interactions, [interaction])

    @patch("server.LLM")
    @patch("server.load_config")
    @patch("server.is_self_evolution_enabled", return_value=False)
    def test_chat_requires_backend_self_evolution_setting(self, setting_mock, config_mock, llm_mock):
        result = AgentResult(
            reply="blocked",
            model_reply="",
            thought="",
            action="official.read_file",
            params={},
            action_result="自我进化未开启：official.read_file",
            action_payload=None,
            response="blocked",
            runtime_nodes=[],
            thread_id="default",
            step_count=0,
        )

        with patch.object(server.agent_runtime, "run", return_value=result) as run_mock:
            server.chat(
                server.ChatRequest(
                    message="读取代码",
                    allowTools=True,
                    allowSkills=True,
                    allowSelfEvolution=True,
                )
            )

        run_mock.assert_called_once()
        self.assertFalse(run_mock.call_args.kwargs["allow_self_evolution"])

    @patch("server.agent_loop")
    def test_legacy_agent_loop_fallback_still_returns_chat_response_shape(self, loop_mock):
        loop_mock.return_value = (
            "Thought: 用户查询天气。\nAction: official.query_weather\nParams: {}",
            "用户查询天气。",
            "official.query_weather",
            {},
            "汕头未来1天天气预报",
            "",
        )

        response = server.run_legacy_agent_loop(FakeLLM(), "汕头天气", True)

        self.assertIn("# ✅ 查询结果", response.reply)
        self.assertIn("汕头未来1天天气预报", response.reply)
        self.assertEqual(response.model_reply, "Thought: 用户查询天气。\nAction: official.query_weather\nParams: {}")
        self.assertEqual(response.thought, "用户查询天气。")
        self.assertEqual(response.action, "official.query_weather")
        self.assertEqual(response.params, {})
        self.assertEqual(response.action_result, "汕头未来1天天气预报")
        self.assertEqual(response.response, "")

    def test_run_agent_runtime_returns_memory_candidate_for_business_info(self):
        result = AgentResult(
            reply="订单已记录。",
            model_reply="Thought: 识别订单。\nAction: none\nParams: {}\nResponse: 订单已记录。",
            thought="识别订单。",
            action="none",
            params={},
            action_result="无操作",
            action_payload=None,
            response="订单已记录。",
            runtime_nodes=[],
            thread_id="default",
            step_count=0,
        )

        with patch.object(server.agent_runtime, "run", return_value=result):
            response = server.run_agent_runtime(FakeLLM(), "客户A 下单 SKU-001 共 20 件", True)

        self.assertIsNotNone(response.memory_candidate)
        self.assertEqual(response.memory_candidate["reason"], "confirmation_required")
        self.assertEqual(len(server.memory.pending_candidates), 1)
        self.assertEqual(server.memory.long_term, [])

    def test_factory_quote_runtime_result_does_not_return_memory_candidate(self):
        result = AgentResult(
            reply="工厂报价查询结果。",
            model_reply="Thought: 查询报价。\nAction: local.factory_quote\nParams: {}",
            thought="查询报价。",
            action="local.factory_quote",
            params={"factory_name": "潘多多", "sku": "PD1102", "quantity": 100},
            action_result="货号：PD1102 数量：100 成本：15.225 报价总额：1674.75",
            action_payload=None,
            response="工厂报价查询结果。",
            runtime_nodes=[],
            thread_id="default",
            step_count=0,
        )

        with patch.object(server.agent_runtime, "run", return_value=result):
            response = server.run_agent_runtime(
                FakeLLM(),
                "潘多多 PD1102产品尺寸、装箱数、毛重是多少？100个按10%利润帮我算报价",
                True,
            )

        self.assertIsNone(response.memory_candidate)
        self.assertEqual(server.memory.pending_candidates, [])

    def test_memory_confirm_and_reject_endpoints(self):
        server.memory.propose_long_term("客户A 下单 SKU-001 共 20 件", "order", module="private_sales")

        confirmed = server.confirm_memory_candidate(server.MemoryConfirmRequest())

        self.assertTrue(confirmed["ok"])
        self.assertEqual(len(server.memory.long_term), 1)

        server.memory.propose_long_term("客户B 下单 SKU-002 共 3 件", "order", module="private_sales")
        rejected = server.reject_memory_candidate(server.MemoryConfirmRequest())

        self.assertTrue(rejected["ok"])
        self.assertEqual(len(server.memory.pending_candidates), 0)
        self.assertEqual(len(server.memory.long_term), 1)

    def test_memory_confirm_returns_404_without_pending_candidate(self):
        with self.assertRaises(HTTPException) as exc:
            server.confirm_memory_candidate(server.MemoryConfirmRequest())

        self.assertEqual(exc.exception.status_code, 404)


if __name__ == "__main__":
    unittest.main()
