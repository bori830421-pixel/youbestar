import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import openpyxl

from tools.excel_feedback_store import save_excel_feedback
from tools.excel_preview_tool import preview_excel, preview_excel_file, save_uploaded_excel
from tools.excel_table_classifier import classify_table, standard_field_catalog


class ExcelPreviewToolTest(unittest.TestCase):
    def make_workbook(self) -> Path:
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "multi_sheet.xlsx"
        workbook = openpyxl.Workbook()

        first = workbook.active
        first.title = "报价表"
        first.append(["说明", "", ""])
        first.append(["货号", "品名", "成本价"])
        for index in range(1, 26):
            first.append([f"SKU-{index:03d}", f"产品{index}", index])

        second = workbook.create_sheet("联系人")
        second.append(["工厂", "业务员", "电话"])
        second.append(["潘多多", "潘小姐", "13502760103"])

        workbook.save(path)
        workbook.close()
        return path

    def test_preview_reads_all_sheets_headers_and_first_twenty_rows(self):
        result = preview_excel_file(self.make_workbook())

        self.assertIs(result["ok"], True)
        self.assertEqual(result["kind"], "excel_preview")
        self.assertEqual(result["title"], "Excel 通用表格识别预览")
        sheets = result["data"]["sheets"]
        self.assertEqual([sheet["name"] for sheet in sheets], ["报价表", "联系人"])
        self.assertEqual(sheets[0]["headers"], ["货号", "品名", "成本价"])
        self.assertEqual(sheets[0]["leading_rows"], [["说明"]])
        self.assertEqual(len(sheets[0]["rows"]), 20)
        self.assertEqual(sheets[0]["rows"][0], ["SKU-001", "产品1", "1"])
        self.assertEqual(sheets[0]["classification"]["status"], "recognized")
        self.assertEqual(sheets[0]["classification"]["category"], "quote")
        self.assertEqual(sheets[0]["classification"]["category_label"], "报价表")
        mapped_labels = {
            mapping["source_header"]: mapping["standard_label"]
            for mapping in sheets[0]["classification"]["field_mappings"]
            if mapping["status"] == "mapped"
        }
        self.assertEqual(mapped_labels["货号"], "商品编码")
        self.assertEqual(mapped_labels["品名"], "商品名称")
        self.assertEqual(mapped_labels["成本价"], "成本单价")
        self.assertEqual(sheets[1]["headers"], ["工厂", "业务员", "电话"])
        self.assertEqual(sheets[1]["rows"][0], ["潘多多", "潘小姐", "13502760103"])
        self.assertEqual(sheets[1]["classification"]["status"], "unknown")
        self.assertIn("未识别工作表", result["summary"])

    def test_preview_skill_accepts_source_path(self):
        path = self.make_workbook()
        result = preview_excel({"source_path": str(path)})

        self.assertIs(result["ok"], True)
        self.assertEqual(result["summary"]["工作表数"], 2)
        self.assertEqual(result["data"]["saved_path"], str(path))

    def test_preview_accepts_directory_with_filename(self):
        path = self.make_workbook()
        result = preview_excel({"source_path": str(path.parent), "filename": path.name})

        self.assertIs(result["ok"], True)
        self.assertEqual(result["data"]["saved_path"], str(path))

    def test_preview_directory_without_filename_lists_choices(self):
        path = self.make_workbook()
        second = path.parent / "other.xlsx"
        workbook = openpyxl.Workbook()
        workbook.save(second)
        workbook.close()

        result = preview_excel({"source_path": str(path.parent)})

        self.assertIs(result["ok"], False)
        self.assertIn("source_path 是文件夹", result["message"])
        self.assertIn(path.name, result["message"])

    def test_uploaded_excel_uses_unique_saved_paths_for_same_filename(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            local_home = Path(temp_dir) / "YoubestarLocal"
            with patch.dict("os.environ", {"YOUBESTAR_LOCAL_HOME": str(local_home)}):
                first = save_uploaded_excel("folder-a/quote.xlsx", BytesIO(b"first"))
                second = save_uploaded_excel("folder-b/quote.xlsx", BytesIO(b"second"))

                self.assertNotEqual(first, second)
                self.assertEqual(first.parent, local_home / "imports")
                self.assertEqual(second.parent, local_home / "imports")
                self.assertTrue(first.name.endswith("_quote.xlsx"))
                self.assertTrue(second.name.endswith("_quote.xlsx"))
                self.assertEqual(first.read_bytes(), b"first")
                self.assertEqual(second.read_bytes(), b"second")

    def test_header_detection_prefers_keyword_row_over_dense_data_row(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "factory_header.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["吉贵（原中意）玩具厂报价表", "", "", "", ""])
        sheet.append(["", "", "", "", ""])
        sheet.append(["联系人：朱召深   联系电话：13923663353", "", "", "", ""])
        sheet.append(["", "货号", "包装/品名", "包装", "厂价(RMB)", "包装规格"])
        sheet.append(["", "8001", "英.阿文.序列棋牌/纸棋盘", "天地盒", 23.3, "27*39.5*5.5"])
        sheet.append(["", "8001A", "英.阿文.序列棋牌/纸棋盘", "飞机盒", 21.5, "27*39.5*5.5"])
        workbook.save(path)
        workbook.close()

        result = preview_excel_file(path, preview_rows=2)
        sheet_preview = result["data"]["sheets"][0]

        self.assertEqual(sheet_preview["header_row"], 4)
        self.assertEqual(sheet_preview["leading_rows"][0], ["吉贵（原中意）玩具厂报价表"])
        self.assertEqual(sheet_preview["leading_rows"][1], ["联系人：朱召深 联系电话：13923663353"])
        self.assertEqual(sheet_preview["headers"][:5], ["未命名列1", "货号", "包装/品名", "包装", "厂价(RMB)"])
        self.assertEqual(sheet_preview["rows"][0][:5], ["", "8001", "英.阿文.序列棋牌/纸棋盘", "天地盒", "23.3"])

    def test_unknown_sheet_does_not_force_business_category(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "unknown.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "待判断"
        sheet.append(["名称", "数量", "备注"])
        sheet.append(["甲", 3, "样例"])
        workbook.save(path)
        workbook.close()

        result = preview_excel_file(path)
        classification = result["data"]["sheets"][0]["classification"]

        self.assertEqual(classification["status"], "unknown")
        self.assertEqual(classification["category_label"], "未识别")
        self.assertTrue(classification["needs_confirmation"])
        name_mapping = classification["field_mappings"][0]
        self.assertEqual(name_mapping["source_header"], "名称")
        self.assertEqual(name_mapping["status"], "pending_confirmation")
        self.assertEqual(name_mapping["legacy_status"], "ambiguous")
        self.assertTrue(name_mapping["needs_confirmation"])
        self.assertGreaterEqual(name_mapping["mapping_score"], 0.65)
        self.assertGreaterEqual(len(name_mapping["candidates"]), 2)

    def test_unknown_field_generates_confirmation_proposal(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "quote_extra.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "报价表"
        sheet.append(["货号", "品名", "成本价", "包装率"])
        sheet.append(["SKU-001", "产品1", 1, "10%"])
        workbook.save(path)
        workbook.close()

        result = preview_excel_file(path)
        classification = result["data"]["sheets"][0]["classification"]

        self.assertEqual(classification["status"], "recognized")
        self.assertEqual(classification["category"], "quote")
        self.assertTrue(classification["needs_confirmation"])
        proposal = classification["change_proposals"][0]
        self.assertEqual(proposal["source_header"], "包装率")
        self.assertEqual(proposal["suggested_field"], "packing_rate")
        self.assertEqual(proposal["standard_label"], "包装率")
        self.assertTrue(proposal["needs_confirmation"])

    def test_new_product_fields_are_standard_mappings(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "product_fields.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "商品资料"
        sheet.append(["货号", "规格编码", "商品简称", "玩具类型", "标准售价", "批发价", "销项税率", "进项税率"])
        sheet.append(["PD1102", "PD1102-RED", "红棋", "益智棋类", 29.9, 18.5, "13%", "6%"])
        workbook.save(path)
        workbook.close()

        result = preview_excel_file(path)
        classification = result["data"]["sheets"][0]["classification"]

        self.assertEqual(classification["status"], "recognized")
        self.assertEqual(classification["category"], "product")
        mapped = {
            mapping["source_header"]: (mapping["standard_field"], mapping["standard_label"])
            for mapping in classification["field_mappings"]
            if mapping["status"] == "mapped"
        }
        self.assertEqual(mapped["规格编码"], ("sku_code", "规格编码"))
        self.assertEqual(mapped["商品简称"], ("short_name", "商品简称"))
        self.assertEqual(mapped["玩具类型"], ("toy_type", "玩具类型"))
        self.assertEqual(mapped["标准售价"], ("standard_price", "标准售价"))
        self.assertEqual(mapped["批发价"], ("wholesale_price", "批发价"))
        self.assertEqual(mapped["销项税率"], ("output_tax_rate", "销项税率"))
        self.assertEqual(mapped["进项税率"], ("input_tax_rate", "进项税率"))

    def test_field_mapping_scores_choose_best_candidate_and_mark_uncertain_dimensions(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "dimension_fields.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "规格字段"
        sheet.append(["货号", "品牌价", "包装规格(cm)", "棋盘规格", "毛/净重(KG)", "单品克重"])
        sheet.append(["PD1102", 18.5, "27*39.5*5.5", "28.9*28.9*2.3", "23/21", "580"])
        workbook.save(path)
        workbook.close()

        result = preview_excel_file(path)
        mappings = {
            mapping["source_header"]: mapping
            for mapping in result["data"]["sheets"][0]["classification"]["field_mappings"]
        }

        self.assertEqual(mappings["品牌价"]["standard_field"], "cost_unit_price")
        self.assertEqual(mappings["品牌价"]["status"], "mapped")
        self.assertGreaterEqual(mappings["品牌价"]["mapping_score"], 0.85)
        self.assertEqual(mappings["包装规格(cm)"]["standard_field"], "product_size_cm")
        self.assertEqual(mappings["包装规格(cm)"]["status"], "pending_confirmation")
        self.assertTrue(mappings["包装规格(cm)"]["needs_confirmation"])
        self.assertGreaterEqual(len(mappings["包装规格(cm)"]["candidates"]), 2)
        self.assertEqual(mappings["棋盘规格"]["standard_field"], "product_size_cm")
        self.assertEqual(mappings["毛/净重(KG)"]["standard_field"], "weight_text")
        self.assertEqual(mappings["单品克重"]["status"], "pending_confirmation")
        self.assertIn(
            "single_gross_weight_g",
            {candidate["standard_field"] for candidate in mappings["单品克重"]["candidates"]},
        )

    def test_field_catalog_contains_dimension_weight_and_spec_fields(self):
        field_codes = {field["code"] for field in standard_field_catalog()}

        self.assertTrue(
            {
                "product_size_cm",
                "package_size_cm",
                "inner_box_quantity",
                "inner_box_size_cm",
                "carton_size_cm",
                "carton_gross_weight_kg",
                "carton_net_weight_kg",
                "single_gross_weight_g",
                "single_net_weight_g",
                "shipping_packaged_weight_g",
                "dimension_text",
                "weight_text",
                "product_spec",
            }.issubset(field_codes)
        )

    def test_low_score_header_is_unknown_with_empty_candidates(self):
        classification = classify_table(["神秘字段XYZ"])
        mapping = classification["field_mappings"][0]

        self.assertEqual(mapping["status"], "unknown")
        self.assertEqual(mapping["legacy_status"], "unmapped")
        self.assertEqual(mapping["mapping_score"], 0.0)
        self.assertEqual(mapping["candidates"], [])
        self.assertEqual(mapping["proposal"]["source_header"], "神秘字段XYZ")

    def test_confirmed_feedback_overrides_category_and_field_mapping(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            local_home = Path(temp_dir) / "YoubestarLocal"
            with patch.dict("os.environ", {"YOUBESTAR_LOCAL_HOME": str(local_home)}):
                headers = ["名称", "数量", "备注"]
                feedback = save_excel_feedback(
                    {
                        "headers": headers,
                        "category": "inventory",
                        "field_mappings": {"名称": "product_name"},
                        "scope": "template",
                    }
                )
                self.assertTrue(feedback["ok"])

                path = Path(temp_dir) / "feedback.xlsx"
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "修正表"
                sheet.append(headers)
                sheet.append(["积木A", 3, "样例"])
                workbook.save(path)
                workbook.close()

                result = preview_excel_file(path)
                classification = result["data"]["sheets"][0]["classification"]
                self.assertEqual(classification["status"], "recognized")
                self.assertEqual(classification["category"], "inventory")
                self.assertEqual(classification["category_label"], "库存表")
                self.assertTrue(classification["feedback"]["applied"])
                first_mapping = classification["field_mappings"][0]
                self.assertEqual(first_mapping["standard_field"], "product_name")
                self.assertEqual(first_mapping["standard_label"], "商品名称")
                self.assertEqual(first_mapping["source"], "user_feedback")


if __name__ == "__main__":
    unittest.main()
